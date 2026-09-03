# -*- coding: utf-8 -*-
"""
DSS 수출입(외자) 관리 프로그램
================================
「1. 수출입.영업.관리 업무 정리」 엑셀의 '업무관련' 시트(영업_외자 업무 흐름 0~8단계)와
'OfferSheet 송부' 시트(신용장 기한 산정 규칙·메일 양식)를 기준으로 만든 업무 관리 프로그램.

  · 데이터는 이 폴더의 data\\dss_trade.db (SQLite) 에만 저장됩니다. (외부 전송 없음)
  · 실행 :  실행.bat   또는   python dss_trade_manager.py
"""

import os
import re
import shutil
import sqlite3
import datetime
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

APP_TITLE = "DSS 수출입(외자) 관리"
APP_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(APP_DIR, "data")
DB_PATH = os.path.join(DATA_DIR, "dss_trade.db")

# ---------------------------------------------------------------------------
# 디자인 토큰
# ---------------------------------------------------------------------------
F = "맑은 고딕"
FONT = (F, 10)
FONT_B = (F, 10, "bold")
FONT_S = (F, 9)
FONT_SB = (F, 9, "bold")
FONT_H = (F, 13, "bold")
FONT_XL = (F, 24, "bold")

C_HEADER = "#1B2F52"      # 상단 바
C_HEADER_SUB = "#93A7C4"
C_SIDE = "#243449"        # 좌측 메뉴
C_SIDE_HOVER = "#31465F"
C_SIDE_ACTIVE = "#0F62FE"
C_SIDE_TEXT = "#C6D2E1"
C_SIDE_TEXT_ON = "#FFFFFF"
C_BG = "#EEF1F5"          # 본문 배경
C_CARD = "#FFFFFF"
C_LINE = "#D5DCE5"
C_TEXT = "#1B2430"
C_MUTED = "#6B7787"

C_BLUE = "#0F62FE"
C_AMBER = "#D9820B"
C_RED = "#D02F35"
C_GREEN = "#1E8E4E"
C_GRAY = "#5A6B7D"

ROW_ODD = "#F7F9FC"
ROW_LATE = "#FDE4E4"
ROW_SOON = "#FEF3DA"
ROW_DONE_FG = "#8E99A6"
ROW_OK = "#E4F4E9"


# ---------------------------------------------------------------------------
# 업무 기준 정보  (엑셀 '업무관련' 시트 — 1. 영업_외자 업무 흐름)
# ---------------------------------------------------------------------------

STAGES = [
    (0, "견적서 요청 접수",
     "국내기업(고객사)로부터 메일 또는 전화로 문의 접수",
     "영업관련업무 엑셀시트"),
    (1, "교산에 견적내용 전송",
     "최종고객사 / 모델 / 수량 / 필요시기 / 일반·예비기 구분을 정리하여 전송.\n"
     "견적서 받으면 프린트하여 보관",
     "견적 폴더"),
    (2, "발주서(P.O) 접수",
     "국내기업(고객사)로부터 발주서 수령 (외자 SCM 사이트에서 확인 및 프린트).\n"
     "업무 엑셀시트(외자관리폴더) 기입 및 NAS 외자발주서 폴더 정리",
     "외자발주서 폴더"),
    (3, "교산 주문 및 납기 관리",
     "교산 제작소에 P.O / Debit Note 작성 후 함께 발송.\n"
     "주문서 수령 사인백을 받아 국내기업에 송부, 발주서와 함께 파일철 정리.\n"
     "제품 생산 및 납기 일정 확인, 수시 Follow-up → 변경사항 실시간 안내",
     "데빗노트 폴더"),
    (4, "신용장(L/C) 개설 요청",
     "교산제작소가 Offer Sheet 송부 → 내용 확인 후 국내기업에 송부.\n"
     "메일로 양식·개설은행·Open 기한 등 구체적 안내.\n"
     "선적 1주일 전까지 개설요청서 수령 → 일본 Confirm → 응답서(L/C번호)를 교산에 송부.\n"
     "수정사항 필요 시 Amend 신청",
     "신용장 폴더"),
    (5, "선적서류 수령",
     "교산제작소에서 Invoice / Packing List / B/L 등 수령.\n"
     "서류 이상유무 검토, 이상 있을 시 교산에 수정 요청",
     "각 발주서와 함께 저장"),
    (6, "선적서류 국내기업 발송",
     "국내기업(구매처) 담당자에게 선적서류 송부. 필요시 원본/사본 구분하여 전달",
     "각 발주서와 함께 저장"),
    (7, "원산지증명서(C/O) 발급",
     "교산에 원산지증명서 발급 요청 후 선적서류와 함께 제출.\n"
     "도착 즉시 국내기업에 전달",
     "원산지증명서 폴더"),
    (8, "커미션 청구 및 수령",
     "모든 납품 완료/검수 후 커미션 청구서를 교산에 송부 (청구서는 Debit Note 에 있음).\n"
     "파일명은 INVOICE DSS<날짜>_Payment Summary for~ 로 시작.\n"
     "청구한 커미션은 매월 말 25일 전후로 은행을 통해 수령",
     "커미션청구서 폴더"),
]
STAGE_COUNT = len(STAGES)
STAGE_NAME = dict((n, nm) for n, nm, _d, _s in STAGES)

STATUSES = ["진행중", "보류", "완료", "취소"]

DEFAULT_MASTER = {
    "고객사": ["기타"],
    "최종고객사": ["기타"],
    "구분": ["일반", "예비기", "케이블", "부품", "기타"],
    "통화": ["JPY", "USD", "KRW", "EUR"],
}
MASTER_CATS = ["고객사", "최종고객사", "구분", "통화"]

DEFAULT_SETTINGS = {
    "company": "㈜디에스에스",
    "user": "",
    "ship_days": "14",        # 기본 Latest shipment = FCA + n일
    "expiry_days": "28",      # 기본 Expiry date    = FCA + n일
    "lc_lead": "7",           # 신용장 개설요청서 수령 기한 = FCA − n일
    "comm_due": "30",         # 커미션 청구 후 n일 경과 시 미수령 경고
    "horizon": "21",          # n일 앞까지 알림에 표시
    # 자료방 폴더 (기본값: 이 프로그램 폴더의 상위에 있는 '자료방')
    "archive_path": os.path.join(os.path.dirname(APP_DIR), "자료방"),
}
NUMERIC_SETTINGS = ("ship_days", "expiry_days", "lc_lead", "comm_due", "horizon")

FILE_KINDS = {
    ".xlsx": "Excel", ".xlsm": "Excel", ".xls": "Excel", ".csv": "CSV",
    ".docx": "Word", ".doc": "Word", ".pptx": "PowerPoint", ".ppt": "PowerPoint",
    ".pdf": "PDF", ".hwp": "한글", ".hwpx": "한글",
    ".msg": "메일", ".eml": "메일", ".txt": "텍스트",
    ".png": "이미지", ".jpg": "이미지", ".jpeg": "이미지", ".gif": "이미지",
    ".bmp": "이미지", ".tif": "이미지", ".tiff": "이미지",
    ".zip": "압축", ".7z": "압축", ".rar": "압축", ".alz": "압축",
    ".dwg": "도면", ".step": "3D", ".stp": "3D", ".lnk": "바로가기",
}


def file_kind(name):
    return FILE_KINDS.get(os.path.splitext(name)[1].lower(), "파일")


def fsize(n):
    if n is None:
        return ""
    if n < 1024:
        return "%d B" % n
    if n < 1024 * 1024:
        return "%.0f KB" % (n / 1024.0)
    if n < 1024 * 1024 * 1024:
        return "%.1f MB" % (n / 1048576.0)
    return "%.1f GB" % (n / 1073741824.0)


def mtime(path):
    try:
        t = datetime.datetime.fromtimestamp(os.path.getmtime(path))
        return t.strftime("%Y-%m-%d %H:%M")
    except OSError:
        return ""


def recycle(paths):
    """파일/폴더를 Windows 휴지통으로 보낸다. (성공 개수, 실패 목록) 반환."""
    paths = [os.path.abspath(p) for p in paths if os.path.exists(p)]
    if not paths:
        return 0, []
    try:
        import ctypes
        from ctypes import wintypes

        class SHFILEOPSTRUCTW(ctypes.Structure):
            _fields_ = [("hwnd", wintypes.HWND),
                        ("wFunc", wintypes.UINT),
                        ("pFrom", wintypes.LPCWSTR),
                        ("pTo", wintypes.LPCWSTR),
                        ("fFlags", ctypes.c_uint16),
                        ("fAnyOperationsAborted", wintypes.BOOL),
                        ("hNameMappings", ctypes.c_void_p),
                        ("lpszProgressTitle", wintypes.LPCWSTR)]

        FO_DELETE = 3
        FOF_SILENT = 0x0004
        FOF_NOCONFIRMATION = 0x0010
        FOF_ALLOWUNDO = 0x0040
        FOF_NOERRORUI = 0x0400

        op = SHFILEOPSTRUCTW()
        op.hwnd = None
        op.wFunc = FO_DELETE
        # pFrom 은 이중 NULL 로 끝나야 한다
        op.pFrom = "\0".join(paths) + "\0\0"
        op.pTo = None
        op.fFlags = FOF_ALLOWUNDO | FOF_NOCONFIRMATION | FOF_SILENT | FOF_NOERRORUI
        rc = ctypes.windll.shell32.SHFileOperationW(ctypes.byref(op))
        if rc == 0 and not op.fAnyOperationsAborted:
            remain = [p for p in paths if os.path.exists(p)]
            return len(paths) - len(remain), remain
    except Exception:
        pass
    return 0, list(paths)


def open_path(path, parent=None):
    """파일/폴더를 Windows 기본 프로그램으로 연다."""
    try:
        os.startfile(path)
        return True
    except Exception as e:
        messagebox.showerror(APP_TITLE, "열 수 없습니다.\n\n%s\n\n%s" % (path, e),
                             parent=parent)
        return False


# ---------------------------------------------------------------------------
# 날짜 유틸
# ---------------------------------------------------------------------------

def norm_date(s):
    """'20261108', '2026.11.8', '2026/11/08' → 'YYYY-MM-DD'"""
    if s is None:
        return ""
    s = str(s).strip().replace(".", "-").replace("/", "-").replace(" ", "")
    if not s:
        return ""
    m = re.match(r"^(\d{4})-?(\d{1,2})-?(\d{1,2})$", s)
    if not m:
        return s
    try:
        return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3))).isoformat()
    except ValueError:
        return s


def parse_date(s):
    s = norm_date(s)
    if not re.match(r"^\d{4}-\d{2}-\d{2}$", s or ""):
        return None
    try:
        return datetime.date(*[int(x) for x in s.split("-")])
    except ValueError:
        return None


def add_days(s, n):
    d = parse_date(s)
    return (d + datetime.timedelta(days=n)).isoformat() if d else ""


def today_str():
    return datetime.date.today().isoformat()


def dday(s):
    d = parse_date(s)
    return None if d is None else (d - datetime.date.today()).days


def dday_text(n):
    if n is None:
        return ""
    return "D-DAY" if n == 0 else ("D-%d" % n if n > 0 else "D+%d" % (-n))


def kdate(s):
    d = parse_date(s)
    return "%d년 %d월 %d일" % (d.year, d.month, d.day) if d else (s or "")


def dwidth(s):
    import unicodedata
    return sum(2 if unicodedata.east_asian_width(ch) in ("W", "F") else 1 for ch in s)


def dpad(s, w):
    return s + " " * max(0, w - dwidth(s))


def money(s):
    """'48000000' → '48,000,000' (숫자가 아니면 원본 유지)"""
    t = str(s or "").replace(",", "").strip()
    if not t:
        return ""
    try:
        return "{:,}".format(int(t))
    except ValueError:
        return str(s)


# ---------------------------------------------------------------------------
# 데이터베이스
# ---------------------------------------------------------------------------

DEAL_COLUMNS = [
    "code", "customer", "customer_pic", "end_user", "site", "model", "qty",
    "kind", "need_date", "po_no", "po_date", "debit_no", "offer_no",
    "fca_date", "latest_shipment", "expiry_date",
    "lc_no", "lc_request_date", "lc_open_date", "lc_bank", "lc_amend",
    "amount", "currency",
    "comm_rate", "comm_amount", "comm_invoice_no", "comm_billed", "comm_received",
    "status", "note", "archive_dir",
]


class Database(object):
    def __init__(self, path):
        d = os.path.dirname(path)
        if not os.path.isdir(d):
            os.makedirs(d)
        fresh = not os.path.exists(path)
        self.con = sqlite3.connect(path)
        self.con.row_factory = sqlite3.Row
        self._create()
        self._seed_master()
        if fresh:
            self._seed_samples()

    # -- 스키마 -------------------------------------------------------------
    def _create(self):
        c = self.con.cursor()
        c.execute("""CREATE TABLE IF NOT EXISTS deal (
            id INTEGER PRIMARY KEY AUTOINCREMENT, code TEXT UNIQUE,
            customer TEXT, customer_pic TEXT, end_user TEXT, site TEXT,
            model TEXT, qty TEXT, kind TEXT, need_date TEXT,
            po_no TEXT, po_date TEXT, debit_no TEXT, offer_no TEXT,
            fca_date TEXT, latest_shipment TEXT, expiry_date TEXT,
            lc_no TEXT, lc_request_date TEXT, lc_open_date TEXT,
            lc_bank TEXT, lc_amend TEXT, amount TEXT, currency TEXT,
            comm_rate TEXT, comm_amount TEXT, comm_invoice_no TEXT,
            comm_billed TEXT, comm_received TEXT,
            status TEXT DEFAULT '진행중', note TEXT, archive_dir TEXT,
            created_at TEXT, updated_at TEXT)""")
        # 이전 버전 DB 보완
        have = set(r[1] for r in c.execute("PRAGMA table_info(deal)").fetchall())
        for col in DEAL_COLUMNS:
            if col not in have:
                c.execute("ALTER TABLE deal ADD COLUMN %s TEXT" % col)
        c.execute("""CREATE TABLE IF NOT EXISTS stage_log (
            deal_id INTEGER NOT NULL, stage_no INTEGER NOT NULL,
            done_date TEXT, memo TEXT, PRIMARY KEY (deal_id, stage_no))""")
        c.execute("""CREATE TABLE IF NOT EXISTS code_master (
            category TEXT NOT NULL, value TEXT NOT NULL, sort INTEGER DEFAULT 0,
            PRIMARY KEY (category, value))""")
        c.execute("""CREATE TABLE IF NOT EXISTS lc_rule (
            customer TEXT PRIMARY KEY, ship_days INTEGER, expiry_days INTEGER)""")
        c.execute("""CREATE TABLE IF NOT EXISTS setting (
            key TEXT PRIMARY KEY, value TEXT)""")
        c.execute("""CREATE TABLE IF NOT EXISTS flow_node (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            seq INTEGER, title TEXT, descr TEXT, tip TEXT,
            store TEXT, folder TEXT, stage_no INTEGER,
            x INTEGER, y INTEGER, shape TEXT, color TEXT)""")
        have = set(r[1] for r in c.execute("PRAGMA table_info(flow_node)").fetchall())
        for col, typ in (("x", "INTEGER"), ("y", "INTEGER"),
                         ("shape", "TEXT"), ("color", "TEXT")):
            if col not in have:
                c.execute("ALTER TABLE flow_node ADD COLUMN %s %s" % (col, typ))
        c.execute("""CREATE TABLE IF NOT EXISTS flow_edge (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            src INTEGER, dst INTEGER, label TEXT)""")
        self.con.commit()

    def _seed_master(self):
        cur = self.con.cursor()
        for k, v in DEFAULT_SETTINGS.items():
            cur.execute("INSERT OR IGNORE INTO setting (key,value) VALUES (?,?)", (k, v))
        if not cur.execute("SELECT 1 FROM code_master LIMIT 1").fetchone():
            for cat, vals in DEFAULT_MASTER.items():
                for i, v in enumerate(vals):
                    cur.execute("INSERT OR IGNORE INTO code_master VALUES (?,?,?)",
                                (cat, v, i))
        if not cur.execute("SELECT 1 FROM flow_node LIMIT 1").fetchone():
            self.reset_flow()
        self.con.commit()

    # -- 업무 플로우 ---------------------------------------------------------
    def reset_flow(self):
        """엑셀 '업무관련' 시트의 0~8단계 + 순서 연결선으로 되돌린다."""
        self.con.execute("DELETE FROM flow_edge")
        self.con.execute("DELETE FROM flow_node")
        ids = []
        for no, name, desc, store in STAGES:
            r, c = divmod(no, 3)
            cur = self.con.execute(
                "INSERT INTO flow_node "
                "(seq,title,descr,tip,store,folder,stage_no,x,y,shape,color) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                (no * 10, "%d. %s" % (no, name), desc, STAGE_TIP.get(no, ""),
                 store, "", no, 60 + c * 300, 50 + r * 160, "box", ""))
            ids.append(cur.lastrowid)
        for a, b in zip(ids, ids[1:]):
            self.con.execute("INSERT INTO flow_edge (src,dst,label) VALUES (?,?,'')",
                             (a, b))
        self.con.commit()

    def flow_nodes(self):
        return self.con.execute("SELECT * FROM flow_node ORDER BY seq, id").fetchall()

    def flow_node(self, nid):
        return self.con.execute("SELECT * FROM flow_node WHERE id=?", (nid,)).fetchone()

    def add_flow(self, data, after_seq=None):
        if after_seq is None:
            seq = (self.con.execute(
                "SELECT COALESCE(MAX(seq),0) FROM flow_node").fetchone()[0] or 0) + 10
        else:
            seq = after_seq + 5
        cur = self.con.execute(
            "INSERT INTO flow_node "
            "(seq,title,descr,tip,store,folder,stage_no,x,y,shape,color) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            (seq, data.get("title", ""), data.get("descr", ""), data.get("tip", ""),
             data.get("store", ""), data.get("folder", ""), data.get("stage_no", -1),
             int(data.get("x", 60)), int(data.get("y", 50)),
             data.get("shape", "box"), data.get("color", "")))
        self.con.commit()
        self._renumber_flow()
        return cur.lastrowid

    def update_flow(self, nid, data):
        self.con.execute(
            "UPDATE flow_node SET title=?, descr=?, tip=?, store=?, folder=?, "
            "stage_no=?, shape=?, color=? WHERE id=?",
            (data.get("title", ""), data.get("descr", ""), data.get("tip", ""),
             data.get("store", ""), data.get("folder", ""), data.get("stage_no", -1),
             data.get("shape", "box"), data.get("color", ""), nid))
        self.con.commit()

    def set_flow_pos(self, nid, x, y):
        self.con.execute("UPDATE flow_node SET x=?, y=? WHERE id=?",
                         (int(x), int(y), nid))
        self.con.commit()

    def delete_flow(self, nid):
        self.con.execute("DELETE FROM flow_edge WHERE src=? OR dst=?", (nid, nid))
        self.con.execute("DELETE FROM flow_node WHERE id=?", (nid,))
        self.con.commit()
        self._renumber_flow()

    def auto_layout(self, per_row=3, x0=60, y0=50, dx=300, dy=160):
        for i, r in enumerate(self.flow_nodes()):
            row, col = divmod(i, per_row)
            self.con.execute("UPDATE flow_node SET x=?, y=? WHERE id=?",
                             (x0 + col * dx, y0 + row * dy, r["id"]))
        self.con.commit()

    def move_flow(self, nid, direction):
        ids = [r["id"] for r in self.flow_nodes()]
        if nid not in ids:
            return
        i = ids.index(nid)
        j = i + direction
        if not (0 <= j < len(ids)):
            return
        ids[i], ids[j] = ids[j], ids[i]
        for k, x in enumerate(ids):
            self.con.execute("UPDATE flow_node SET seq=? WHERE id=?", (k * 10, x))
        self.con.commit()

    def _renumber_flow(self):
        for k, r in enumerate(self.flow_nodes()):
            self.con.execute("UPDATE flow_node SET seq=? WHERE id=?", (k * 10, r["id"]))
        self.con.commit()

    # -- 연결선 -------------------------------------------------------------
    def flow_edges(self):
        return self.con.execute("SELECT * FROM flow_edge ORDER BY id").fetchall()

    def add_edge(self, src, dst, label=""):
        if src == dst:
            return None
        dup = self.con.execute("SELECT id FROM flow_edge WHERE src=? AND dst=?",
                               (src, dst)).fetchone()
        if dup:
            return dup["id"]
        cur = self.con.execute("INSERT INTO flow_edge (src,dst,label) VALUES (?,?,?)",
                               (src, dst, label))
        self.con.commit()
        return cur.lastrowid

    def delete_edge(self, eid):
        self.con.execute("DELETE FROM flow_edge WHERE id=?", (eid,))
        self.con.commit()

    def set_edge_label(self, eid, label):
        self.con.execute("UPDATE flow_edge SET label=? WHERE id=?", (label, eid))
        self.con.commit()

    def _seed_samples(self):
        y = datetime.date.today().year
        a = self.insert({
            "code": "%d-001" % y, "customer": "샘플고객사A", "customer_pic": "담당자A",
            "end_user": "엔드유저A", "site": "사이트A", "model": "샘플 장비 60kW",
            "qty": "27CH", "kind": "일반", "po_no": "PO-SAMPLE-001",
            "po_date": add_days(today_str(), -30), "offer_no": "0000-00001",
            "debit_no": "DN-SAMPLE-001", "fca_date": add_days(today_str(), 12),
            "amount": "48000000", "currency": "JPY", "comm_rate": "5",
            "lc_bank": "샘플은행", "status": "진행중",
            "note": "[샘플 데이터] 확인 후 삭제하세요."})
        for n, dt in ((0, -40), (1, -37), (2, -30), (3, -25)):
            self.set_stage(a, n, add_days(today_str(), dt), "")
        self.recalc_lc(a)

        b = self.insert({
            "code": "%d-002" % y, "customer": "샘플고객사B", "customer_pic": "담당자B",
            "end_user": "엔드유저B", "site": "사이트B", "model": "샘플 장비 20kW",
            "qty": "3SET", "kind": "일반", "po_no": "PO-SAMPLE-002",
            "po_date": add_days(today_str(), -70), "offer_no": "0000-00002",
            "fca_date": add_days(today_str(), -20), "lc_no": "LC-SAMPLE-0001",
            "lc_bank": "샘플은행", "amount": "21500000", "currency": "JPY",
            "comm_rate": "5", "comm_amount": "1075000",
            "comm_invoice_no": "INVOICE-SAMPLE-001", "comm_billed": add_days(today_str(), -35),
            "status": "진행중", "note": "[샘플 데이터] 확인 후 삭제하세요."})
        for n, dt in ((0, -90), (1, -88), (2, -70), (3, -60),
                      (4, -35), (5, -22), (6, -20), (7, -18)):
            self.set_stage(b, n, add_days(today_str(), dt), "")
        self.recalc_lc(b)

    # -- 설정 / 기준정보 -----------------------------------------------------
    def get_setting(self, key, default=""):
        r = self.con.execute("SELECT value FROM setting WHERE key=?", (key,)).fetchone()
        return r["value"] if r else default

    def get_int(self, key, default=0):
        try:
            return int(str(self.get_setting(key, default)).strip())
        except ValueError:
            return default

    def set_setting(self, key, value):
        self.con.execute("INSERT OR REPLACE INTO setting (key,value) VALUES (?,?)",
                         (key, str(value)))
        self.con.commit()

    def codes(self, category):
        return [r["value"] for r in self.con.execute(
            "SELECT value FROM code_master WHERE category=? ORDER BY sort, value",
            (category,)).fetchall()]

    def add_code(self, category, value):
        n = self.con.execute("SELECT COALESCE(MAX(sort),-1)+1 FROM code_master "
                             "WHERE category=?", (category,)).fetchone()[0]
        self.con.execute("INSERT OR IGNORE INTO code_master VALUES (?,?,?)",
                         (category, value, n))
        self.con.commit()

    def del_code(self, category, value):
        self.con.execute("DELETE FROM code_master WHERE category=? AND value=?",
                         (category, value))
        self.con.commit()

    def lc_rules(self):
        return self.con.execute(
            "SELECT customer, ship_days, expiry_days FROM lc_rule ORDER BY customer"
        ).fetchall()

    def set_lc_rule(self, customer, ship_days, expiry_days):
        self.con.execute("INSERT OR REPLACE INTO lc_rule VALUES (?,?,?)",
                         (customer, int(ship_days), int(expiry_days)))
        self.con.commit()

    def del_lc_rule(self, customer):
        self.con.execute("DELETE FROM lc_rule WHERE customer=?", (customer,))
        self.con.commit()

    def lc_rule(self, customer):
        """(선적기한 가산일, 유효기간 가산일). 고객사 전용 규칙이 없으면 기본값."""
        r = self.con.execute(
            "SELECT ship_days, expiry_days FROM lc_rule WHERE customer=?",
            (customer or "",)).fetchone()
        if r:
            return int(r["ship_days"]), int(r["expiry_days"])
        return self.get_int("ship_days", 14), self.get_int("expiry_days", 28)

    # -- 안건 ---------------------------------------------------------------
    def all_deals(self):
        return self.con.execute("SELECT * FROM deal ORDER BY code DESC").fetchall()

    def get(self, deal_id):
        return self.con.execute("SELECT * FROM deal WHERE id=?", (deal_id,)).fetchone()

    def stages(self, deal_id):
        rows = self.con.execute("SELECT stage_no, done_date, memo FROM stage_log "
                                "WHERE deal_id=?", (deal_id,)).fetchall()
        return dict((r["stage_no"], (r["done_date"] or "", r["memo"] or "")) for r in rows)

    def done_set(self, deal_id):
        return set(n for n, (d, _m) in self.stages(deal_id).items() if d)

    def next_code(self):
        y = datetime.date.today().year
        mx = 0
        for r in self.con.execute("SELECT code FROM deal WHERE code LIKE ?",
                                  ("%d-%%" % y,)).fetchall():
            m = re.match(r"^%d-(\d+)$" % y, r["code"] or "")
            if m:
                mx = max(mx, int(m.group(1)))
        return "%d-%03d" % (y, mx + 1)

    def insert(self, data):
        d = dict((c, "") for c in DEAL_COLUMNS)
        d.update(data)
        d["created_at"] = d["updated_at"] = today_str()
        cols = DEAL_COLUMNS + ["created_at", "updated_at"]
        cur = self.con.execute(
            "INSERT INTO deal (%s) VALUES (%s)" % (",".join(cols), ",".join("?" * len(cols))),
            [d.get(c, "") for c in cols])
        self.con.commit()
        return cur.lastrowid

    def update(self, deal_id, data):
        data = dict(data)
        data["updated_at"] = today_str()
        cols = [c for c in DEAL_COLUMNS if c in data] + ["updated_at"]
        self.con.execute("UPDATE deal SET %s WHERE id=?"
                         % ",".join("%s=?" % c for c in cols),
                         [data.get(c, "") for c in cols] + [deal_id])
        self.con.commit()

    def delete(self, deal_id):
        self.con.execute("DELETE FROM stage_log WHERE deal_id=?", (deal_id,))
        self.con.execute("DELETE FROM deal WHERE id=?", (deal_id,))
        self.con.commit()

    def set_stage(self, deal_id, stage_no, done_date, memo):
        self.con.execute("INSERT OR REPLACE INTO stage_log "
                         "(deal_id, stage_no, done_date, memo) VALUES (?,?,?,?)",
                         (deal_id, stage_no, done_date, memo))
        self.con.commit()

    def recalc_lc(self, deal_id):
        r = self.get(deal_id)
        if not r or not parse_date(r["fca_date"]):
            return
        ship, exp = self.lc_rule(r["customer"])
        self.update(deal_id, {"latest_shipment": add_days(r["fca_date"], ship),
                              "expiry_date": add_days(r["fca_date"], exp)})

    def duplicate(self, deal_id):
        r = self.get(deal_id)
        if not r:
            return None
        data = dict((c, r[c]) for c in DEAL_COLUMNS)
        data["code"] = self.next_code()
        data["status"] = "진행중"
        for k in ("po_no", "comm_billed", "comm_received", "lc_no",
                  "lc_open_date", "comm_invoice_no"):
            data[k] = ""
        return self.insert(data)

    # -- 파생 정보 -----------------------------------------------------------
    def progress(self, deal_id):
        """(완료 단계 수, 다음 단계 번호 or None)"""
        done = self.done_set(deal_id)
        nxt = next((n for n in range(STAGE_COUNT) if n not in done), None)
        return len(done), nxt

    def next_deadline(self, r, done=None):
        """다가오는 기한 (라벨, 날짜, 남은일수). 없으면 ('', '', None)"""
        if done is None:
            done = self.done_set(r["id"])
        lead = self.get_int("lc_lead", 7)
        due = self.get_int("comm_due", 30)
        cand = []
        if 4 not in done and parse_date(r["fca_date"]):
            cand.append(("L/C 요청", add_days(r["fca_date"], -lead)))
        if 5 not in done and parse_date(r["fca_date"]):
            cand.append(("FCA", r["fca_date"]))
        if 5 not in done and parse_date(r["latest_shipment"]):
            cand.append(("선적기한", r["latest_shipment"]))
        if 6 not in done and parse_date(r["expiry_date"]):
            cand.append(("L/C 만료", r["expiry_date"]))
        if 8 not in done and parse_date(r["comm_billed"]) and not parse_date(r["comm_received"]):
            cand.append(("커미션 입금", add_days(r["comm_billed"], due)))
        cand = [c for c in cand if parse_date(c[1])]
        if not cand:
            return "", "", None
        cand.sort(key=lambda c: parse_date(c[1]))
        for label, d in cand:
            n = dday(d)
            if n is not None and n >= 0:
                return label, norm_date(d), n
        label, d = cand[-1]
        return label, norm_date(d), dday(d)


# ---------------------------------------------------------------------------
# 알림
# ---------------------------------------------------------------------------

def compute_alerts(db):
    out = []
    lead = db.get_int("lc_lead", 7)
    due = db.get_int("comm_due", 30)
    horizon = db.get_int("horizon", 21)

    def push(level, deal, item, date_s, detail):
        n = dday(date_s)
        out.append({"level": level, "id": deal["id"], "code": deal["code"] or "",
                    "customer": deal["customer"] or "", "po": deal["po_no"] or "",
                    "item": item, "date": norm_date(date_s), "dday": n,
                    "sort": n if n is not None else 9999, "detail": detail})

    def lv(n):
        return "지연" if n < 0 else ("임박" if n <= 7 else "예정")

    for d in db.all_deals():
        if (d["status"] or "") in ("완료", "취소"):
            continue
        done = db.done_set(d["id"])

        if 4 not in done and parse_date(d["fca_date"]):
            t = add_days(d["fca_date"], -lead)
            n = dday(t)
            if n is not None and n <= horizon:
                push(lv(n), d, "L/C 개설요청 기한", t,
                     "FCA %s 기준 %d일 전까지 개설요청서 수령 필요"
                     % (norm_date(d["fca_date"]), lead))

        if 5 not in done and parse_date(d["fca_date"]):
            n = dday(d["fca_date"])
            if n is not None and n <= horizon:
                push(lv(n), d, "FCA (선적예정)", d["fca_date"],
                     "선적서류 수령 준비 / 납기 Follow-up")

        if 5 not in done and parse_date(d["latest_shipment"]):
            n = dday(d["latest_shipment"])
            if n is not None and n <= horizon:
                push(lv(n), d, "Latest shipment", d["latest_shipment"],
                     "신용장 선적기한 — 초과 시 Amend 필요")

        if 6 not in done and parse_date(d["expiry_date"]):
            n = dday(d["expiry_date"])
            if n is not None and n <= horizon:
                push(lv(n), d, "L/C 유효기간 만료", d["expiry_date"],
                     "선적서류 매입 / 발송 완료 확인")

        if 7 in done and 8 not in done and not parse_date(d["comm_billed"]):
            push("확인", d, "커미션 미청구", "",
                 "원산지증명서까지 완료 — 커미션 청구서 송부 필요")

        b = parse_date(d["comm_billed"])
        if b and not parse_date(d["comm_received"]):
            passed = (datetime.date.today() - b).days
            if passed >= due:
                push("지연", d, "커미션 미수령", add_days(d["comm_billed"], due),
                     "%s 청구 후 %d일 경과 — 매월 25일 전후 입금 확인"
                     % (norm_date(d["comm_billed"]), passed))

    rank = {"지연": 0, "임박": 1, "확인": 2, "예정": 3}
    out.sort(key=lambda a: (rank.get(a["level"], 9), a["sort"]))
    return out


# ---------------------------------------------------------------------------
# 공통 위젯
# ---------------------------------------------------------------------------

class Card(tk.Frame):
    """대시보드 요약 카드."""

    def __init__(self, parent, title, color, on_click=None):
        tk.Frame.__init__(self, parent, bg=C_CARD, highlightthickness=1,
                          highlightbackground=C_LINE, cursor="hand2" if on_click else "")
        tk.Frame(self, bg=color, height=4).pack(fill="x")
        body = tk.Frame(self, bg=C_CARD)
        body.pack(fill="both", expand=True, padx=18, pady=(12, 14))
        self.v_num = tk.StringVar(value="0")
        self.v_sub = tk.StringVar(value="")
        tk.Label(body, text=title, font=FONT_SB, bg=C_CARD, fg=C_MUTED).pack(anchor="w")
        tk.Label(body, textvariable=self.v_num, font=FONT_XL, bg=C_CARD,
                 fg=color).pack(anchor="w", pady=(2, 0))
        tk.Label(body, textvariable=self.v_sub, font=FONT_S, bg=C_CARD,
                 fg=C_MUTED).pack(anchor="w")
        if on_click:
            for w in [self, body] + list(body.winfo_children()):
                w.bind("<Button-1>", lambda e: on_click())

    def set(self, num, sub=""):
        self.v_num.set(str(num))
        self.v_sub.set(sub)


class Grid(ttk.Frame):
    """정렬·줄무늬가 붙은 Treeview 래퍼."""

    def __init__(self, parent, columns, on_open=None, height=12, selectmode="extended"):
        ttk.Frame.__init__(self, parent)
        self.spec = columns                      # [(key, title, width, anchor)]
        self.on_open = on_open
        self._sort = (None, False)

        keys = [c[0] for c in columns]
        self.tv = ttk.Treeview(self, columns=keys, show="headings",
                               height=height, selectmode=selectmode)
        for key, title, w, anchor in columns:
            self.tv.heading(key, text=title,
                            command=lambda k=key: self._sort_by(k))
            self.tv.column(key, width=w, anchor=anchor, stretch=(w >= 200))
        vs = ttk.Scrollbar(self, orient="vertical", command=self.tv.yview)
        hs = ttk.Scrollbar(self, orient="horizontal", command=self.tv.xview)
        self.tv.configure(yscroll=vs.set, xscroll=hs.set)
        self.tv.grid(row=0, column=0, sticky="nsew")
        vs.grid(row=0, column=1, sticky="ns")
        hs.grid(row=1, column=0, sticky="ew")
        self.rowconfigure(0, weight=1)
        self.columnconfigure(0, weight=1)

        self.tv.tag_configure("odd", background=ROW_ODD)
        self.tv.tag_configure("late", background=ROW_LATE)
        self.tv.tag_configure("soon", background=ROW_SOON)
        self.tv.tag_configure("ok", background=ROW_OK)
        self.tv.tag_configure("done", foreground=ROW_DONE_FG)
        if on_open:
            self.tv.bind("<Double-1>", lambda e: self._open())
            self.tv.bind("<Return>", lambda e: self._open())

    def _open(self):
        sel = self.tv.selection()
        if sel and self.on_open:
            self.on_open(sel[0])

    def set_rows(self, rows):
        """rows: [(iid, [values...], tag)]"""
        self._rows = rows
        self._render()

    def _render(self):
        key, rev = self._sort
        rows = list(getattr(self, "_rows", []))
        if key:
            i = [c[0] for c in self.spec].index(key)
            rows.sort(key=lambda r: str(r[1][i]), reverse=rev)
        sel = set(self.tv.selection())
        for x in self.tv.get_children():
            self.tv.delete(x)
        for n, (iid, values, tag) in enumerate(rows):
            tags = [t for t in (tag,) if t]
            if not tags and n % 2:
                tags = ["odd"]
            self.tv.insert("", "end", iid=iid, values=values, tags=tuple(tags))
        for s in sel:
            if self.tv.exists(s):
                self.tv.selection_add(s)

    def _sort_by(self, key):
        k, rev = self._sort
        self._sort = (key, not rev if k == key else False)
        for c_key, title, _w, _a in self.spec:
            mark = ""
            if c_key == self._sort[0]:
                mark = "  ▼" if self._sort[1] else "  ▲"
            self.tv.heading(c_key, text=title + mark)
        self._render()

    def selection(self):
        return self.tv.selection()

    def select(self, iid):
        if self.tv.exists(iid):
            self.tv.selection_set(iid)
            self.tv.focus(iid)
            self.tv.see(iid)


def toolbar(parent):
    bar = tk.Frame(parent, bg=C_BG)
    bar.pack(fill="x", pady=(0, 10))
    return bar


def page_title(parent, title, desc=""):
    box = tk.Frame(parent, bg=C_BG)
    box.pack(fill="x", pady=(0, 12))
    tk.Label(box, text=title, font=FONT_H, bg=C_BG, fg=C_TEXT).pack(anchor="w")
    if desc:
        tk.Label(box, text=desc, font=FONT_S, bg=C_BG, fg=C_MUTED).pack(anchor="w", pady=(2, 0))
    return box


# ---------------------------------------------------------------------------
# 폼 정의
# ---------------------------------------------------------------------------
# (key, 라벨, 위젯, 옵션)   위젯: entry / date / combo / combo_free
FORM_TABS = [
    ("기본정보", [
        ("code",         "안건번호",         "entry",      {"readonly": True}),
        ("status",       "상태",             "combo",      {"master": None, "values": STATUSES}),
        ("customer",     "고객사(국내기업)",  "combo_free", {"master": "고객사"}),
        ("customer_pic", "고객사 담당자",     "entry",      {}),
        ("end_user",     "최종고객사",        "combo_free", {"master": "최종고객사"}),
        ("site",         "사이트 / 지역",     "entry",      {}),
        ("model",        "모델 / 품명",       "entry",      {}),
        ("qty",          "수량",             "entry",      {}),
        ("kind",         "구분",             "combo_free", {"master": "구분"}),
        ("need_date",    "필요시기",          "date",       {}),
    ]),
    ("발주 · 납기", [
        ("po_no",    "발주서 P.O 번호",   "entry", {}),
        ("po_date",  "P.O 접수일",        "date",  {}),
        ("debit_no", "Debit Note 번호",   "entry", {}),
        ("offer_no", "Offer Sheet 번호",  "entry", {"hint": "공급사 내부번호 (예: 0000-00000)"}),
        ("fca_date", "FCA (선적예정일)",  "date",  {"hint": "이 날짜로 신용장 기한이 자동 계산됩니다"}),
        ("amount",   "금액",              "entry", {}),
        ("currency", "통화",              "combo", {"master": "통화"}),
    ]),
    ("신용장(L/C)", [
        ("lc_request_date", "개설요청 안내일",  "date",  {}),
        ("lc_bank",         "개설은행",        "entry", {}),
        ("lc_no",           "L/C 번호",        "entry", {}),
        ("lc_open_date",    "개설(응답서)일",   "date",  {}),
        ("latest_shipment", "Latest shipment", "date",  {"hint": "선적기한"}),
        ("expiry_date",     "Expiry date",     "date",  {"hint": "신용장 유효기간"}),
        ("lc_amend",        "Amend 이력",      "entry", {}),
    ]),
    ("커미션", [
        ("comm_rate",       "커미션 요율(%)",  "entry", {}),
        ("comm_amount",     "커미션 금액",     "entry", {}),
        ("comm_invoice_no", "청구서 번호",     "entry",
         {"hint": "INVOICE DSS<날짜>_Payment Summary for ~"}),
        ("comm_billed",     "청구일",         "date",  {}),
        ("comm_received",   "수령일",         "date",  {"hint": "매월 말 25일 전후 입금"}),
    ]),
]
FORM_KEYS = [f[0] for t in FORM_TABS for f in t[1]]
DATE_KEYS = set(f[0] for t in FORM_TABS for f in t[1] if f[2] == "date")
FORM_META = dict((f[0], (f[1], f[2], f[3])) for t in FORM_TABS for f in t[1])

# 단계별로 입력해야 하는 항목
STAGE_FIELDS = {
    0: ["customer", "customer_pic", "end_user", "site", "model", "qty", "kind", "need_date"],
    1: ["amount", "currency"],
    2: ["po_no", "po_date", "amount", "currency"],
    3: ["debit_no", "offer_no", "fca_date"],
    4: ["lc_request_date", "lc_bank", "lc_no", "lc_open_date",
        "latest_shipment", "expiry_date", "lc_amend"],
    5: [],
    6: [],
    7: [],
    8: ["comm_rate", "comm_amount", "comm_invoice_no", "comm_billed", "comm_received"],
}
# 단계별 자료방 폴더 (프로젝트 폴더 하위 이름 / 자료방 최상위 분류 키워드)
STAGE_FOLDER = {
    0: ("02_견적", "견적"), 1: ("02_견적", "견적"),
    2: ("03_PO", "수주"), 3: ("04_Offer sheet", "발주"),
    4: ("05_신용장", "계약"), 5: ("06_선적", "선적"),
    6: ("06_선적", "선적"), 7: ("06_선적", "통관"),
    8: ("07_정산", "계약"),
}
# 단계별 체크포인트 (놓치기 쉬운 것)
STAGE_TIP = {
    0: "최종고객사 · 모델 · 수량 · 필요시기 · 일반/예비기 구분을 빠짐없이 확인하세요.",
    1: "교산에 보낸 내용과 받은 견적서를 자료방에 함께 남겨두세요.",
    2: "외자 SCM 사이트에서 P.O 를 확인·프린트하고 발주서 폴더에 정리하세요.",
    3: "FCA 를 입력하면 신용장 기한이 자동 계산됩니다. 사인백 수령 여부를 확인하세요.",
    4: "개설요청서는 선적 1주일 전까지 받아야 합니다. 응답서(L/C번호)는 교산에 송부하세요.",
    5: "Invoice / Packing List / B/L 이상 유무를 검토하고, 이상 시 교산에 수정 요청하세요.",
    6: "원본/사본 구분이 필요한지 고객사 담당자에게 확인하세요.",
    7: "원산지증명서는 도착 즉시 국내기업에 전달합니다.",
    8: "청구서 파일명은 INVOICE DSS<날짜>_Payment Summary for~ 로 시작합니다.",
}


# ===========================================================================
# 페이지
# ===========================================================================

class Page(tk.Frame):
    def __init__(self, app, parent=None):
        tk.Frame.__init__(self, parent or app.body, bg=C_BG)
        self.app = app
        self.db = app.db
        self.inner = tk.Frame(self, bg=C_BG)
        self.inner.pack(fill="both", expand=True, padx=22, pady=18)

    def refresh(self):
        pass


# --------------------------------------------------------------- 대시보드
class DashboardPage(Page):
    def __init__(self, app):
        Page.__init__(self, app)
        page_title(self.inner, "대시보드",
                   "오늘 기준으로 챙겨야 할 항목입니다. 목록을 더블클릭하면 해당 안건이 열립니다.")

        cards = tk.Frame(self.inner, bg=C_BG)
        cards.pack(fill="x", pady=(0, 16))
        self.c_active = Card(cards, "진행중 안건", C_BLUE,
                             lambda: app.open_deals_view("all"))
        self.c_soon = Card(cards, "7일 내 임박", C_AMBER, self.refresh)
        self.c_late = Card(cards, "기한 경과", C_RED, self.refresh)
        self.c_comm = Card(cards, "커미션 미수령", C_GREEN,
                           lambda: app.open_deals_view("comm"))
        for i, c in enumerate((self.c_active, self.c_soon, self.c_late, self.c_comm)):
            c.grid(row=0, column=i, sticky="ew", padx=(0, 12))
            cards.columnconfigure(i, weight=1)

        wrap = tk.Frame(self.inner, bg=C_BG)
        wrap.pack(fill="both", expand=True)
        wrap.columnconfigure(0, weight=3)
        wrap.columnconfigure(1, weight=2)
        wrap.rowconfigure(0, weight=1)

        left = tk.LabelFrame(wrap, text=" 기한 알림 ", bg=C_BG, fg=C_TEXT,
                             font=FONT_B, bd=0)
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 12))
        self.g_alert = Grid(left, [
            ("level", "구분", 60, "center"), ("dday", "D-Day", 70, "center"),
            ("date", "기준일", 96, "center"), ("item", "항목", 150, "w"),
            ("code", "안건", 80, "center"), ("customer", "고객사", 110, "w"),
            ("detail", "내용", 300, "w")], on_open=self._open, height=14)
        self.g_alert.pack(fill="both", expand=True, pady=(6, 0))

        right = tk.LabelFrame(wrap, text=" 고객사별 진행 현황 ", bg=C_BG, fg=C_TEXT,
                              font=FONT_B, bd=0)
        right.grid(row=0, column=1, sticky="nsew")
        self.g_cust = Grid(right, [
            ("customer", "고객사", 130, "w"), ("cnt", "진행중", 60, "center"),
            ("late", "지연", 55, "center"), ("soon", "임박", 55, "center"),
            ("amt", "금액 합계", 130, "e")], height=14, selectmode="browse")
        self.g_cust.pack(fill="both", expand=True, pady=(6, 0))

    def _open(self, iid):
        self.app.open_deal(int(iid.split("_")[1]))

    def refresh(self):
        alerts = compute_alerts(self.db)
        rows = self.db.all_deals()
        active = [r for r in rows if (r["status"] or "진행중") == "진행중"]
        late = [a for a in alerts if a["level"] == "지연"]
        soon = [a for a in alerts if a["level"] == "임박"]
        comm = [r for r in rows if parse_date(r["comm_billed"])
                and not parse_date(r["comm_received"])]

        self.c_active.set(len(active), "전체 %d건 중" % len(rows))
        self.c_soon.set(len(soon), "7일 이내 처리 필요")
        self.c_late.set(len(late), "기한이 지났습니다" if late else "없음")
        self.c_comm.set(len(comm), "청구 후 입금 대기")

        self.g_alert.set_rows([
            ("a%d_%d" % (i, a["id"]),
             [a["level"], dday_text(a["dday"]), a["date"], a["item"],
              a["code"], a["customer"], a["detail"]],
             {"지연": "late", "임박": "soon", "확인": "ok"}.get(a["level"], ""))
            for i, a in enumerate(alerts)])

        by = {}
        for r in active:
            k = r["customer"] or "(미지정)"
            b = by.setdefault(k, {"cnt": 0, "amt": 0, "late": 0, "soon": 0})
            b["cnt"] += 1
            try:
                b["amt"] += int(str(r["amount"] or "0").replace(",", "") or 0)
            except ValueError:
                pass
        for a in alerts:
            k = a["customer"] or "(미지정)"
            if k in by:
                if a["level"] == "지연":
                    by[k]["late"] += 1
                elif a["level"] == "임박":
                    by[k]["soon"] += 1
        self.g_cust.set_rows([
            ("c%d" % i, [k, v["cnt"], v["late"] or "", v["soon"] or "",
                         money(v["amt"]) if v["amt"] else ""],
             "late" if v["late"] else ("soon" if v["soon"] else ""))
            for i, (k, v) in enumerate(sorted(by.items(), key=lambda x: -x[1]["cnt"]))])


# ------------------------------------------------------------- 안건 관리
class DealsPage(Page):
    def __init__(self, app):
        Page.__init__(self, app)
        self.current_id = None
        self.dirty = False
        self._loading = False
        self.vars = {}
        self.widgets = {}
        self.combos = []
        self.stage_vars = {}
        self.cur_step = 0
        self._last_fca = ""

        self.view = "all"
        page_title(self.inner, "안건 관리",
                   "위 목록에서 안건을 고르면 아래에 현재 단계가 열립니다. "
                   "0 → 8 단계를 순서대로 진행하고, 보기를 바꾸면 단계별 현황을 볼 수 있습니다.")

        bar = toolbar(self.inner)
        self._btn(bar, "＋ 신규", self.new_deal, primary=True)
        self._btn(bar, "저장", self.save_deal, primary=True)
        self._btn(bar, "복제", self.duplicate_deal)
        self._btn(bar, "삭제", self.delete_deal)
        tk.Frame(bar, bg=C_LINE, width=1).pack(side="left", fill="y", padx=10, pady=3)
        self._btn(bar, "Offer Sheet 메일", self.make_mail)
        self._btn(bar, "엑셀 내보내기", app.export_excel)

        tk.Label(bar, text="검색", bg=C_BG, fg=C_MUTED, font=FONT_S).pack(side="left", padx=(16, 4))
        self.v_search = tk.StringVar()
        e = ttk.Entry(bar, textvariable=self.v_search, width=20, font=FONT)
        e.pack(side="left")
        e.bind("<KeyRelease>", lambda ev: self.refresh_grid())

        tk.Label(bar, text="고객사", bg=C_BG, fg=C_MUTED, font=FONT_S).pack(side="left", padx=(12, 4))
        self.v_fcust = tk.StringVar(value="전체")
        self.cb_cust = ttk.Combobox(bar, textvariable=self.v_fcust, width=14,
                                    state="readonly", font=FONT)
        self.cb_cust.pack(side="left")
        self.cb_cust.bind("<<ComboboxSelected>>", lambda ev: self.refresh_grid())

        tk.Label(bar, text="상태", bg=C_BG, fg=C_MUTED, font=FONT_S).pack(side="left", padx=(12, 4))
        self.v_fstat = tk.StringVar(value="진행중")
        cb = ttk.Combobox(bar, textvariable=self.v_fstat, width=8, state="readonly",
                          values=["전체"] + STATUSES, font=FONT)
        cb.pack(side="left")
        cb.bind("<<ComboboxSelected>>", lambda ev: self.refresh_grid())

        # 보기 전환 (전체 / 발주·납기 / 신용장 / 선적·서류 / 커미션)
        vbar = tk.Frame(self.inner, bg=C_BG)
        vbar.pack(fill="x", pady=(0, 6))
        tk.Label(vbar, text="보기", bg=C_BG, fg=C_MUTED, font=FONT_S
                 ).pack(side="left", padx=(0, 8))
        self.view_tabs = {}
        for key, label, _c, _r, _s in VIEWS:
            b = tk.Label(vbar, text="  %s  " % label, font=FONT, bg="#FFFFFF",
                         fg=C_TEXT, padx=8, pady=5, cursor="hand2",
                         highlightthickness=1, highlightbackground=C_LINE)
            b.pack(side="left", padx=(0, 5))
            b.bind("<Button-1>", lambda e, k=key: self.show_view(k))
            self.view_tabs[key] = b
        self.v_sum = tk.StringVar()
        tk.Label(vbar, textvariable=self.v_sum, bg=C_BG, fg=C_TEXT,
                 font=FONT_B).pack(side="right")

        pane = ttk.PanedWindow(self.inner, orient="vertical")
        pane.pack(fill="both", expand=True)

        top = tk.Frame(pane, bg=C_BG)
        pane.add(top, weight=3)
        self.grids = {}
        for key, _l, cols, _r, _s in VIEWS:
            g = Grid(top, cols, height=7)
            g.tv.bind("<<TreeviewSelect>>", self.on_select)
            self.grids[key] = g
        self.grid_deals = self.grids["all"]
        self.grid_deals.pack(fill="both", expand=True)

        bottom = tk.Frame(pane, bg=C_BG)
        pane.add(bottom, weight=7)
        self._build_detail(bottom)
        self.paint_view_tabs()

    def _btn(self, bar, text, cmd, primary=False):
        b = tk.Button(bar, text=text, command=cmd, font=FONT_B if primary else FONT,
                      bg=C_BLUE if primary else "#FFFFFF",
                      fg="#FFFFFF" if primary else C_TEXT,
                      activebackground="#0B4FCB" if primary else "#E8EDF3",
                      activeforeground="#FFFFFF" if primary else C_TEXT,
                      relief="flat", bd=0, padx=14, pady=6, cursor="hand2",
                      highlightthickness=1, highlightbackground=C_LINE)
        b.pack(side="left", padx=(0, 6))
        return b

    # -- 상세 폼 ------------------------------------------------------------
    def _scroll_area(self, parent, bg=C_CARD):
        cv = tk.Canvas(parent, highlightthickness=0, bg=bg)
        sb = ttk.Scrollbar(parent, orient="vertical", command=cv.yview)
        inner = tk.Frame(cv, bg=bg)
        inner.bind("<Configure>", lambda e: cv.configure(scrollregion=cv.bbox("all")))
        win = cv.create_window((0, 0), window=inner, anchor="nw")
        cv.bind("<Configure>", lambda e: cv.itemconfig(win, width=e.width))
        cv.configure(yscrollcommand=sb.set)
        cv.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        cv.bind("<Enter>", lambda e: cv.bind_all(
            "<MouseWheel>", lambda ev: cv.yview_scroll(int(-ev.delta / 120), "units")))
        cv.bind("<Leave>", lambda e: cv.unbind_all("<MouseWheel>"))
        return inner

    def _field_row(self, parent, key, row, bg=C_CARD, lw=17):
        """FORM_META 기준으로 라벨+입력칸 한 줄. 같은 key 는 StringVar 를 공유한다."""
        label, kind, opt = FORM_META[key]
        tk.Label(parent, text=label, font=FONT, bg=bg, fg=C_TEXT,
                 width=lw, anchor="w").grid(row=row, column=0, sticky="w",
                                            padx=(0, 12), pady=5)
        if key not in self.vars:
            v = tk.StringVar()
            v.trace("w", lambda *a: self.mark_dirty())
            self.vars[key] = v
        var = self.vars[key]

        holder = tk.Frame(parent, bg=bg)
        holder.grid(row=row, column=1, sticky="ew", pady=5)
        holder.columnconfigure(0, weight=1)

        if kind in ("combo", "combo_free"):
            w = ttk.Combobox(holder, textvariable=var, font=FONT,
                             state="readonly" if kind == "combo" else "normal",
                             values=opt.get("values", []))
            if opt.get("master"):
                self.combos.append((w, opt["master"]))
        else:
            w = ttk.Entry(holder, textvariable=var, font=FONT)
            if opt.get("readonly"):
                w.configure(state="readonly")
        w.grid(row=0, column=0, sticky="ew")
        self.widgets.setdefault(key, w)

        if kind == "date":
            w.bind("<FocusOut>", lambda e, k=key: self.fix_date(k))
            tk.Button(holder, text="오늘", font=FONT_S, relief="flat", bd=0,
                      bg="#EEF2F7", fg=C_TEXT, padx=8, pady=2, cursor="hand2",
                      command=lambda k=key: self.vars[k].set(today_str())
                      ).grid(row=0, column=1, padx=(5, 0))
            if key == "fca_date":
                tk.Button(holder, text="기한 자동계산", command=self.autocalc_lc,
                          font=FONT_S, relief="flat", bd=0, bg="#E3ECFB", fg=C_BLUE,
                          padx=10, pady=2, cursor="hand2").grid(row=0, column=2, padx=(5, 0))

        if opt.get("hint"):
            tk.Label(parent, text=opt["hint"], font=FONT_S, bg=bg, fg=C_MUTED
                     ).grid(row=row, column=2, sticky="w", padx=(14, 0))

    def _build_detail(self, parent):
        head = tk.Frame(parent, bg=C_BG)
        head.pack(fill="x", pady=(10, 4))
        self.v_title = tk.StringVar(value="안건을 선택하거나 [＋ 신규] 를 누르세요")
        tk.Label(head, textvariable=self.v_title, font=FONT_B, bg=C_BG,
                 fg=C_TEXT).pack(side="left")
        self.v_prog = tk.StringVar(value="")
        tk.Label(head, textvariable=self.v_prog, font=FONT_S, bg=C_BG,
                 fg=C_MUTED).pack(side="right")

        # 이 안건의 자료방 폴더
        afr = tk.Frame(parent, bg=C_BG)
        afr.pack(fill="x", pady=(0, 6))
        tk.Label(afr, text="자료 폴더", font=FONT_S, bg=C_BG, fg=C_MUTED,
                 width=9, anchor="w").pack(side="left")
        self.vars["archive_dir"] = tk.StringVar()
        self.vars["archive_dir"].trace("w", lambda *a: self.mark_dirty())
        ttk.Entry(afr, textvariable=self.vars["archive_dir"], font=FONT_S
                  ).pack(side="left", fill="x", expand=True)
        for text, cmd in (("폴더 지정", self.pick_archive_dir),
                          ("열기", lambda: self.open_stage_folder(None))):
            tk.Button(afr, text=text, command=cmd, font=FONT_S, relief="flat", bd=0,
                      bg="#FFFFFF", fg=C_TEXT, padx=10, pady=2, cursor="hand2",
                      highlightthickness=1, highlightbackground=C_LINE
                      ).pack(side="left", padx=(5, 0))

        self.nb = ttk.Notebook(parent)
        self.nb.pack(fill="both", expand=True)

        self.tab_flow = tk.Frame(self.nb, bg=C_CARD)
        self.nb.add(self.tab_flow, text="  ▶ 단계별 진행  ")
        self.tab_all = tk.Frame(self.nb, bg=C_CARD)
        self.nb.add(self.tab_all, text="  전체 항목  ")
        self.tab_note = tk.Frame(self.nb, bg=C_CARD)
        self.nb.add(self.tab_note, text="  비고  ")

        self._build_all(self.tab_all)      # StringVar 를 먼저 만든다
        self._build_flow(self.tab_flow)

        self.txt_note = tk.Text(self.tab_note, font=FONT, wrap="word", relief="flat",
                                bg="#FCFDFE", padx=12, pady=10,
                                highlightthickness=1, highlightbackground=C_LINE)
        self.txt_note.pack(fill="both", expand=True, padx=16, pady=14)
        self.txt_note.bind("<KeyRelease>", lambda e: self.mark_dirty())

    def _build_all(self, parent):
        inner = self._scroll_area(parent)
        inner.columnconfigure(1, minsize=250)
        inner.columnconfigure(2, weight=1)
        pad = tk.Frame(inner, bg=C_CARD, width=18)
        pad.grid(row=0, column=3)
        row = 0
        for title, fields in FORM_TABS:
            tk.Label(inner, text=title, font=FONT_B, bg=C_CARD, fg=C_BLUE
                     ).grid(row=row, column=0, columnspan=3, sticky="w",
                            padx=(18, 0), pady=(14, 2))
            row += 1
            for key, _l, _k, _o in fields:
                f = tk.Frame(inner, bg=C_CARD)
                f.grid(row=row, column=0, columnspan=3, sticky="ew", padx=(18, 0))
                f.columnconfigure(1, minsize=250)
                f.columnconfigure(2, weight=1)
                self._field_row(f, key, 0)
                row += 1

    # -- 단계별 진행 ---------------------------------------------------------
    def _build_flow(self, parent):
        strip = tk.Frame(parent, bg=C_CARD)
        strip.pack(fill="x", padx=14, pady=(12, 8))
        self.step_chips = {}
        for no, name, _d, _s in STAGES:
            c = tk.Frame(strip, bg="#EDF1F6", highlightthickness=1,
                         highlightbackground=C_LINE, cursor="hand2")
            c.pack(side="left", fill="x", expand=True, padx=2)
            n = tk.Label(c, text=str(no), font=(F, 12, "bold"), bg="#EDF1F6", fg=C_MUTED)
            n.pack(pady=(5, 0))
            t = tk.Label(c, text=name.split("(")[0].strip()[:8], font=(F, 8),
                         bg="#EDF1F6", fg=C_MUTED)
            t.pack(pady=(0, 5))
            for wdg in (c, n, t):
                wdg.bind("<Button-1>", lambda e, k=no: self.show_step(k))
            self.step_chips[no] = (c, n, t)

        body = tk.Frame(parent, bg=C_CARD)
        body.pack(fill="both", expand=True, padx=6)
        self.step_host = self._scroll_area(body)
        self.step_frames = {}
        for no, name, desc, store in STAGES:
            self.step_frames[no] = self._build_step(self.step_host, no, name, desc, store)

        nav = tk.Frame(parent, bg=C_CARD)
        nav.pack(fill="x", padx=16, pady=(6, 12))
        tk.Button(nav, text="◀ 이전 단계", command=self.prev_step, font=FONT,
                  relief="flat", bd=0, bg="#FFFFFF", fg=C_TEXT, padx=14, pady=7,
                  cursor="hand2", highlightthickness=1, highlightbackground=C_LINE
                  ).pack(side="left")
        tk.Button(nav, text="이 단계 완료하고 다음 ▶", command=self.complete_step,
                  font=FONT_B, relief="flat", bd=0, bg=C_GREEN, fg="#FFFFFF",
                  padx=18, pady=7, cursor="hand2").pack(side="right")
        tk.Button(nav, text="저장", command=self.save_deal, font=FONT_B, relief="flat",
                  bd=0, bg=C_BLUE, fg="#FFFFFF", padx=18, pady=7, cursor="hand2"
                  ).pack(side="right", padx=(0, 8))
        tk.Button(nav, text="다음 단계 ▶", command=self.next_step, font=FONT,
                  relief="flat", bd=0, bg="#FFFFFF", fg=C_TEXT, padx=14, pady=7,
                  cursor="hand2", highlightthickness=1, highlightbackground=C_LINE
                  ).pack(side="right", padx=(0, 8))

    def _build_step(self, parent, no, name, desc, store):
        fr = tk.Frame(parent, bg=C_CARD)
        fr.columnconfigure(0, weight=1)

        card = tk.Frame(fr, bg="#F4F7FB", highlightthickness=1,
                        highlightbackground=C_LINE)
        card.grid(row=0, column=0, sticky="ew", padx=14, pady=(4, 10))
        card.columnconfigure(0, weight=1)
        tk.Label(card, text="%d단계.  %s" % (no, name), font=(F, 12, "bold"),
                 bg="#F4F7FB", fg=C_HEADER, anchor="w"
                 ).grid(row=0, column=0, sticky="w", padx=14, pady=(10, 4))
        tk.Label(card, text=desc, font=FONT_S, bg="#F4F7FB", fg=C_TEXT,
                 justify="left", anchor="w").grid(row=1, column=0, sticky="w",
                                                  padx=14, pady=(0, 6))
        tk.Label(card, text="✔  " + STAGE_TIP[no], font=FONT_S, bg="#F4F7FB",
                 fg=C_AMBER, justify="left", wraplength=760, anchor="w"
                 ).grid(row=2, column=0, sticky="w", padx=14, pady=(0, 6))
        srow = tk.Frame(card, bg="#F4F7FB")
        srow.grid(row=3, column=0, sticky="w", padx=14, pady=(0, 10))
        tk.Label(srow, text="저장소 : %s" % store, font=FONT_S, bg="#F4F7FB",
                 fg=C_MUTED).pack(side="left")
        tk.Button(srow, text="📁 자료 폴더 열기", font=FONT_S, relief="flat", bd=0,
                  bg="#FFFFFF", fg=C_BLUE, padx=10, pady=2, cursor="hand2",
                  highlightthickness=1, highlightbackground=C_LINE,
                  command=lambda k=no: self.open_stage_folder(k)).pack(side="left", padx=(12, 0))

        # 완료 표시
        done = tk.Frame(fr, bg=C_CARD)
        done.grid(row=1, column=0, sticky="ew", padx=14)
        done.columnconfigure(4, weight=1)
        chk = tk.BooleanVar()
        dat = tk.StringVar()
        memo = tk.StringVar()
        for v in (chk, dat, memo):
            v.trace("w", lambda *a: self.mark_dirty())
        self.stage_vars[no] = (chk, dat, memo)
        tk.Checkbutton(done, text="이 단계 완료", variable=chk, font=FONT_B, bg=C_CARD,
                       fg=C_GREEN, activebackground=C_CARD,
                       command=lambda n=no: self.on_stage_check(n)
                       ).grid(row=0, column=0, sticky="w")
        tk.Label(done, text="완료일", font=FONT_S, bg=C_CARD, fg=C_MUTED
                 ).grid(row=0, column=1, padx=(14, 4))
        de = ttk.Entry(done, textvariable=dat, width=12, font=FONT)
        de.grid(row=0, column=2)
        de.bind("<FocusOut>", lambda ev, v=dat: v.set(norm_date(v.get())))
        tk.Label(done, text="메모", font=FONT_S, bg=C_CARD, fg=C_MUTED
                 ).grid(row=0, column=3, padx=(14, 4))
        ttk.Entry(done, textvariable=memo, font=FONT).grid(row=0, column=4, sticky="ew")

        keys = STAGE_FIELDS.get(no, [])
        if keys:
            ttk.Separator(fr, orient="horizontal").grid(row=2, column=0, sticky="ew",
                                                        padx=14, pady=10)
            g = tk.Frame(fr, bg=C_CARD)
            g.grid(row=3, column=0, sticky="ew", padx=14)
            g.columnconfigure(1, minsize=250)
            g.columnconfigure(2, weight=1)
            for i, key in enumerate(keys):
                self._field_row(g, key, i)
        else:
            tk.Label(fr, text="이 단계는 별도 입력 항목이 없습니다. "
                             "완료 시 체크하고 필요하면 메모를 남기세요.",
                     font=FONT_S, bg=C_CARD, fg=C_MUTED
                     ).grid(row=2, column=0, sticky="w", padx=14, pady=(12, 0))
        tk.Frame(fr, bg=C_CARD, height=12).grid(row=4, column=0)
        return fr

    def show_step(self, no):
        if not (0 <= no < STAGE_COUNT):
            return
        for f in self.step_frames.values():
            f.pack_forget()
        self.step_frames[no].pack(fill="both", expand=True)
        self.cur_step = no
        self.paint_steps()
        try:
            self.nb.select(self.tab_flow)
        except tk.TclError:
            pass

    def paint_steps(self):
        """단계 칩 색칠 : 완료=초록 · 현재=파랑 · 미진행=회색"""
        for no, (c, n, t) in self.step_chips.items():
            chk = self.stage_vars[no][0].get()
            if no == getattr(self, "cur_step", 0):
                bg, fg = C_BLUE, "#FFFFFF"
            elif chk:
                bg, fg = "#DCF0E4", C_GREEN
            else:
                bg, fg = "#EDF1F6", C_MUTED
            mark = "✔" if chk and no != getattr(self, "cur_step", 0) else str(no)
            c.configure(bg=bg)
            n.configure(bg=bg, fg=fg, text=mark)
            t.configure(bg=bg, fg=fg)

    def prev_step(self):
        self.show_step(getattr(self, "cur_step", 0) - 1)

    def next_step(self):
        self.show_step(getattr(self, "cur_step", 0) + 1)

    def complete_step(self):
        if not self.current_id:
            messagebox.showinfo(APP_TITLE, "먼저 안건을 선택하거나 [＋ 신규] 를 누르세요.",
                                parent=self)
            return
        no = getattr(self, "cur_step", 0)
        chk, dat, _m = self.stage_vars[no]
        chk.set(True)
        if not dat.get().strip():
            dat.set(today_str())
        self.save_deal()
        if no + 1 < STAGE_COUNT:
            self.show_step(no + 1)
            self.app.set_status("%d단계 완료 → %d. %s" % (no, no + 1, STAGE_NAME[no + 1]))
        else:
            self.app.set_status("마지막 단계까지 완료했습니다.")
            if messagebox.askyesno(APP_TITLE,
                                   "모든 단계가 끝났습니다.\n이 안건의 상태를 '완료' 로 바꿀까요?",
                                   parent=self):
                self.vars["status"].set("완료")
                self.save_deal()

    def goto_current_step(self):
        """진행 중인(첫 미완료) 단계로 이동."""
        nxt = next((n for n in range(STAGE_COUNT)
                    if not self.stage_vars[n][0].get()), STAGE_COUNT - 1)
        self.show_step(nxt)

    # -- 자료 폴더 -----------------------------------------------------------
    def archive_root(self):
        return self.db.get_setting("archive_path", DEFAULT_SETTINGS["archive_path"])

    def pick_archive_dir(self):
        root = self.archive_root()
        cur = self.vars["archive_dir"].get().strip()
        start = os.path.join(root, cur) if cur else root
        p = filedialog.askdirectory(parent=self, title="이 안건의 자료 폴더 선택",
                                    initialdir=start if os.path.isdir(start) else root)
        if not p:
            return
        p = os.path.normpath(p)
        try:
            rel = os.path.relpath(p, root)
        except ValueError:
            rel = p
        self.vars["archive_dir"].set(p if rel.startswith("..") else rel)

    def stage_folder(self, no):
        """단계에 해당하는 자료방 폴더 경로를 찾는다."""
        root = self.archive_root()
        rel = self.vars["archive_dir"].get().strip()
        base = rel if os.path.isabs(rel) else os.path.join(root, rel) if rel else ""
        if no is None:
            return base if os.path.isdir(base) else (root if os.path.isdir(root) else None)
        sub, keyword = STAGE_FOLDER.get(no, (None, None))
        if base and os.path.isdir(base):
            if sub:
                for n in sorted(os.listdir(base)):
                    if n.lower() == sub.lower() and os.path.isdir(os.path.join(base, n)):
                        return os.path.join(base, n)
            return base
        if os.path.isdir(root) and keyword:
            for n in sorted(os.listdir(root)):
                if keyword in n and os.path.isdir(os.path.join(root, n)):
                    return os.path.join(root, n)
        return root if os.path.isdir(root) else None

    def open_stage_folder(self, no):
        path = self.stage_folder(no)
        if not path:
            messagebox.showinfo(APP_TITLE,
                                "자료방 폴더를 찾을 수 없습니다.\n"
                                "메뉴 [도구 → 설정] 에서 자료방 경로를 확인하세요.", parent=self)
            return
        self.app.show("archive")
        self.app.pages["archive"].goto(path)

    # -- 데이터 -------------------------------------------------------------
    def refresh_choices(self):
        self.cb_cust.configure(values=["전체"] + self.db.codes("고객사"))
        for w, cat in self.combos:
            try:
                w.configure(values=self.db.codes(cat))
            except tk.TclError:
                pass

    def refresh(self):
        self.refresh_choices()
        self.refresh_grid()

    # -- 보기 전환 -----------------------------------------------------------
    def paint_view_tabs(self):
        for key, b in self.view_tabs.items():
            on = (key == self.view)
            b.configure(bg=C_BLUE if on else "#FFFFFF",
                        fg="#FFFFFF" if on else C_TEXT,
                        font=FONT_B if on else FONT)

    def show_view(self, key):
        if key not in self.grids:
            return
        self.view = key
        for g in self.grids.values():
            g.pack_forget()
        self.grid_deals = self.grids[key]
        self.grid_deals.pack(fill="both", expand=True)
        self.paint_view_tabs()
        self.refresh_grid()

    def _filtered(self):
        q = (self.v_search.get() or "").strip().lower()
        fc, fs = self.v_fcust.get(), self.v_fstat.get()
        for r in self.db.all_deals():
            if fc != "전체" and (r["customer"] or "") != fc:
                continue
            if fs != "전체" and (r["status"] or "진행중") != fs:
                continue
            if q and q not in " ".join(str(r[c] or "") for c in DEAL_COLUMNS).lower():
                continue
            yield r

    def refresh_grid(self):
        spec = dict((v[0], v) for v in VIEWS)[self.view]
        _k, _l, _c, rowfn, sumfn = spec
        rows = []
        for r in self._filtered():
            values, tag = rowfn(self.db, r)
            rows.append((str(r["id"]), values, tag))
        self.grid_deals.set_rows(rows)
        if self.current_id:
            self.grid_deals.select(str(self.current_id))
        self.v_sum.set(sumfn(rows))
        self.app.set_status("안건 %d건 표시" % len(rows))

    def on_select(self, event=None):
        sel = self.grid_deals.selection()
        if not sel:
            return
        did = int(sel[0])
        if did == self.current_id:
            return
        if not self.confirm_discard():
            if self.current_id:
                self.grid_deals.select(str(self.current_id))
            return
        self.load(did)

    def load(self, deal_id):
        r = self.db.get(deal_id)
        if not r:
            return
        self._loading = True
        self.current_id = deal_id
        for k in FORM_KEYS + ["archive_dir"]:
            self.vars[k].set(r[k] if r[k] is not None else "")
        self.txt_note.delete("1.0", "end")
        self.txt_note.insert("1.0", r["note"] or "")
        st = self.db.stages(deal_id)
        for no, (chk, dat, memo) in self.stage_vars.items():
            d, m = st.get(no, ("", ""))
            chk.set(bool(d))
            dat.set(d)
            memo.set(m)
        n_done, nxt = self.db.progress(deal_id)
        self.v_title.set("[%s]   %s   ·   %s" % (
            r["code"] or "-", r["customer"] or "고객사 미입력", r["model"] or "품명 미입력"))
        self.v_prog.set("진행 %d / %d 단계    ·    현재: %s" % (
            n_done, STAGE_COUNT,
            "완료" if nxt is None else "%d. %s" % (nxt, STAGE_NAME[nxt])))
        self._last_fca = norm_date(r["fca_date"])
        self._loading = False
        self.dirty = False
        self.goto_current_step()

    def clear(self):
        self._loading = True
        for k in FORM_KEYS + ["archive_dir"]:
            self.vars[k].set("")
        self.txt_note.delete("1.0", "end")
        for chk, dat, memo in self.stage_vars.values():
            chk.set(False)
            dat.set("")
            memo.set("")
        self.v_title.set("안건을 선택하거나 [＋ 신규] 를 누르세요")
        self.v_prog.set("")
        self._last_fca = ""
        self._loading = False
        self.dirty = False
        self.show_step(0)

    # -- 동작 ---------------------------------------------------------------
    def mark_dirty(self, *a):
        if not self._loading:
            self.dirty = True

    def fix_date(self, key):
        self.vars[key].set(norm_date(self.vars[key].get()))
        if key == "fca_date":
            self.sync_lc_dates()

    def sync_lc_dates(self):
        """FCA 가 바뀌면 신용장 기한을 따라가게 한다.
        직접 손댄 값은 함부로 덮지 않고 물어본다."""
        fca = self.vars["fca_date"].get().strip()
        if not parse_date(fca):
            return
        ship, exp = self.db.lc_rule(self.vars["customer"].get())
        new_s, new_e = add_days(fca, ship), add_days(fca, exp)
        cur_s = self.vars["latest_shipment"].get().strip()
        cur_e = self.vars["expiry_date"].get().strip()
        old = getattr(self, "_last_fca", "")

        auto = not cur_s and not cur_e
        if not auto and parse_date(old):
            # 이전 FCA 기준 자동계산값 그대로면 사용자가 손대지 않은 것
            auto = (cur_s == add_days(old, ship) and cur_e == add_days(old, exp))

        if auto:
            self.vars["latest_shipment"].set(new_s)
            self.vars["expiry_date"].set(new_e)
        elif (cur_s != new_s or cur_e != new_e) and not self._loading:
            if messagebox.askyesno(
                    APP_TITLE,
                    "FCA 가 %s 로 바뀌었습니다.\n신용장 기한을 다시 계산할까요?\n\n"
                    "Latest shipment :  %s  →  %s\n"
                    "Expiry date       :  %s  →  %s"
                    % (fca, cur_s or "(없음)", new_s, cur_e or "(없음)", new_e),
                    parent=self):
                self.vars["latest_shipment"].set(new_s)
                self.vars["expiry_date"].set(new_e)
        self._last_fca = fca

    def autocalc_lc(self):
        fca = norm_date(self.vars["fca_date"].get())
        if not parse_date(fca):
            messagebox.showinfo(APP_TITLE, "먼저 FCA(선적예정일)를 입력하세요.", parent=self)
            return
        cust = self.vars["customer"].get()
        ship, exp = self.db.lc_rule(cust)
        lead = self.db.get_int("lc_lead", 7)
        self.vars["fca_date"].set(fca)
        self.vars["latest_shipment"].set(add_days(fca, ship))
        self.vars["expiry_date"].set(add_days(fca, exp))
        base = "%s 전용 규칙" % cust if self.db.con.execute(
            "SELECT 1 FROM lc_rule WHERE customer=?", (cust,)).fetchone() else "기본 규칙"
        messagebox.showinfo(APP_TITLE,
                            "%s (+%d일 / +%d일) 적용\n\n"
                            "FCA                : %s\n"
                            "Latest shipment : %s\n"
                            "Expiry date       : %s\n\n"
                            "· 신용장 개설요청서는 %s 까지 수령해야 합니다."
                            % (base, ship, exp, fca, add_days(fca, ship),
                               add_days(fca, exp), add_days(fca, -lead)), parent=self)

    def on_stage_check(self, no):
        chk, dat, _m = self.stage_vars[no]
        if chk.get() and not dat.get().strip():
            dat.set(today_str())
        elif not chk.get():
            dat.set("")
        self.paint_steps()

    def collect(self):
        d = {}
        for k in FORM_KEYS + ["archive_dir"]:
            v = self.vars[k].get().strip()
            if k in DATE_KEYS:
                v = norm_date(v)
            d[k] = v
        d["amount"] = money(d["amount"])
        d["comm_amount"] = money(d["comm_amount"])
        d["note"] = self.txt_note.get("1.0", "end").strip()
        return d

    def new_deal(self):
        if not self.confirm_discard():
            return
        did = self.db.insert({"code": self.db.next_code(), "status": "진행중",
                              "currency": "JPY", "kind": "일반"})
        self.v_fstat.set("진행중")
        self.v_fcust.set("전체")
        self.v_search.set("")
        self.current_id = did
        self.refresh_grid()
        self.grid_deals.select(str(did))
        self.load(did)
        try:
            self.widgets["customer"].focus_set()
        except Exception:
            pass
        self.app.set_status("새 안건 생성됨")

    def save_deal(self):
        if not self.current_id:
            messagebox.showinfo(APP_TITLE, "저장할 안건이 없습니다. [＋ 신규] 를 눌러주세요.",
                                parent=self)
            return
        data = self.collect()
        self.db.update(self.current_id, data)
        for no, (chk, dat, memo) in self.stage_vars.items():
            d = norm_date(dat.get()) if chk.get() else ""
            if chk.get() and not d:
                d = today_str()
            self.db.set_stage(self.current_id, no, d, memo.get().strip())
        self.dirty = False
        cur = self.current_id
        self.refresh_grid()
        self.load(cur)
        self.app.refresh_others("deals")
        self.app.set_status("저장 완료 — %s" % (data.get("code") or ""))

    def duplicate_deal(self):
        if not self.current_id or not self.confirm_discard():
            return
        nid = self.db.duplicate(self.current_id)
        if nid:
            self.current_id = nid
            self.refresh_grid()
            self.grid_deals.select(str(nid))
            self.load(nid)
            self.app.set_status("안건을 복제했습니다.")

    def delete_deal(self):
        sel = self.grid_deals.selection()
        if not sel:
            messagebox.showinfo(APP_TITLE, "삭제할 안건을 목록에서 선택하세요.", parent=self)
            return
        codes = [self.grid_deals.tv.set(i, "code") for i in sel]
        if not messagebox.askyesno(APP_TITLE,
                                   "%d건을 삭제합니다.\n\n%s\n\n되돌릴 수 없습니다. 진행할까요?"
                                   % (len(sel), ", ".join(codes)), parent=self):
            return
        for i in sel:
            self.db.delete(int(i))
        self.current_id = None
        self.clear()
        self.refresh_grid()
        self.app.refresh_others("deals")
        self.app.set_status("%d건 삭제됨" % len(sel))

    def confirm_discard(self):
        if not self.dirty or not self.current_id:
            return True
        a = messagebox.askyesnocancel(APP_TITLE,
                                      "저장하지 않은 변경사항이 있습니다. 저장할까요?", parent=self)
        if a is None:
            return False
        if a:
            self.save_deal()
        else:
            self.dirty = False
        return True

    def open_deal(self, deal_id):
        if not self.confirm_discard():
            return
        self.v_fstat.set("전체")
        self.v_fcust.set("전체")
        self.v_search.set("")
        self.current_id = deal_id
        self.refresh_grid()
        self.grid_deals.select(str(deal_id))
        self.load(deal_id)

    # -- Offer Sheet 메일 ----------------------------------------------------
    def make_mail(self):
        sel = self.grid_deals.selection()
        if not sel:
            messagebox.showinfo(APP_TITLE,
                                "메일에 넣을 안건을 목록에서 선택하세요.\n(Ctrl 키로 여러 건 선택 가능)",
                                parent=self)
            return
        deals = [d for d in (self.db.get(int(i)) for i in sel) if d]
        if not deals:
            return
        first = deals[0]
        lead = self.db.get_int("lc_lead", 7)
        fca = parse_date(first["fca_date"])
        fca_txt = "%d월 %d일" % (fca.month, fca.day) if fca else "____"

        L = ["수신 : %s / %s" % (first["customer"] or "고객사명",
                                first["customer_pic"] or "담당자명"),
             "발신 : %s / %s" % (self.db.get_setting("company", "㈜디에스에스"),
                                self.db.get_setting("user", "")),
             "", "1. 귀사의 일익 번창하심을 기원합니다.", "",
             "2. 교산 선적 전 신용장 개설 관련입니다. (FCA %s)" % fca_txt,
             "   첨부한 Offer sheet 참고하시고 하기의 내용을 반영하시어",
             "   FCA일정 일주일 전까지 신용장 개설 부탁드립니다.", ""]

        hdr = ["NO", "PO", "Offer sheet", "FCA", "Latest shipment", "Expiry date"]
        rows = []
        for i, d in enumerate(deals, 1):
            ship, exp = self.db.lc_rule(d["customer"])
            ls = norm_date(d["latest_shipment"]) or add_days(d["fca_date"], ship)
            ex = norm_date(d["expiry_date"]) or add_days(d["fca_date"], exp)
            rows.append([str(i), d["po_no"] or "", d["offer_no"] or "",
                         norm_date(d["fca_date"]),
                         "%s( 선적일 기준 +%d주)" % (kdate(ls), ship // 7) if ls else "",
                         "%s ( 선적일 기준 +%d주)" % (kdate(ex), exp // 7) if ex else ""])
        w = [max([dwidth(hdr[c])] + [dwidth(r[c]) for r in rows]) for c in range(len(hdr))]

        def fmt(cells):
            return "  " + " | ".join(dpad(cells[c], w[c]) for c in range(len(hdr)))

        L.append(fmt(hdr))
        L.append("  " + "-+-".join("-" * x for x in w))
        for r in rows:
            L.append(fmt(r))
        L += ["", "   ※ 개설은행 : %s" % (first["lc_bank"] or "(기재)"),
              "   ※ 신용장 개설요청서는 %s 까지 회신 부탁드립니다."
              % (add_days(first["fca_date"], -lead) or "____"),
              "", "감사합니다."]
        self.app.show_text_window("Offer Sheet 송부 메일 문안", "\n".join(L))


# ===========================================================================
# 안건 목록 보기 (전체 / 발주·납기 / 신용장 / 선적·서류 / 커미션)
#   각 함수는 안건 1건을 받아 (값 목록, 색 태그) 를 돌려준다.
# ===========================================================================

def _tag_by_days(days):
    if days is None:
        return ""
    return "late" if days < 0 else ("soon" if days <= 7 else "")


def _amount_sum(rows, idx):
    t = 0
    for _i, v, _g in rows:
        try:
            t += int(str(v[idx]).replace(",", "") or 0)
        except ValueError:
            pass
    return t


COLS_ALL = [("code", "안건번호", 84, "center"), ("customer", "고객사", 118, "w"),
            ("end_user", "최종고객사", 92, "w"), ("model", "모델 / 품명", 200, "w"),
            ("qty", "수량", 70, "center"), ("po", "P.O 번호", 116, "w"),
            ("stage", "진행단계", 190, "w"), ("prog", "진척", 52, "center"),
            ("fca", "FCA", 92, "center"), ("next", "다음 기한", 150, "w"),
            ("status", "상태", 60, "center")]


def row_all(db, r):
    n_done, nxt = db.progress(r["id"])
    stage_txt = "완료" if nxt is None else "%d. %s" % (nxt, STAGE_NAME[nxt])
    label, _d, days = db.next_deadline(r)
    tag = "done" if (r["status"] or "") in ("완료", "취소") else _tag_by_days(days)
    return ([r["code"] or "", r["customer"] or "", r["end_user"] or "",
             r["model"] or "", r["qty"] or "", r["po_no"] or "", stage_txt,
             "%d/%d" % (n_done, STAGE_COUNT), norm_date(r["fca_date"]),
             "%s %s" % (label, dday_text(days)) if label else "",
             r["status"] or "진행중"], tag)


def sum_all(rows):
    return "%d건" % len(rows)


COLS_ORDER = [("code", "안건번호", 84, "center"), ("customer", "고객사", 118, "w"),
              ("model", "모델 / 품명", 190, "w"), ("qty", "수량", 66, "center"),
              ("po", "P.O 번호", 116, "w"), ("po_date", "P.O 접수", 90, "center"),
              ("debit", "Debit Note", 105, "w"), ("offer", "Offer Sheet", 98, "w"),
              ("fca", "FCA", 90, "center"), ("dday", "D-Day", 66, "center"),
              ("s2", "2. P.O 접수", 92, "center"), ("s3", "3. 교산 주문", 92, "center"),
              ("amount", "금액", 108, "e"), ("cur", "통화", 46, "center")]


def row_order(db, r):
    st = db.stages(r["id"])
    n = dday(r["fca_date"])
    tag = "" if 5 in db.done_set(r["id"]) else _tag_by_days(n)
    return ([r["code"] or "", r["customer"] or "", r["model"] or "", r["qty"] or "",
             r["po_no"] or "", norm_date(r["po_date"]), r["debit_no"] or "",
             r["offer_no"] or "", norm_date(r["fca_date"]), dday_text(n),
             st.get(2, ("", ""))[0] or "―", st.get(3, ("", ""))[0] or "―",
             money(r["amount"]), r["currency"] or ""], tag)


def sum_order(rows):
    return "%d건    금액 합계 %s" % (len(rows), money(_amount_sum(rows, 12)))


COLS_LC = [("code", "안건번호", 84, "center"), ("customer", "고객사", 118, "w"),
           ("po", "P.O 번호", 116, "w"), ("fca", "FCA", 90, "center"),
           ("req_due", "개설요청 기한", 100, "center"), ("req_dday", "D-Day", 66, "center"),
           ("bank", "개설은행", 88, "w"), ("lc_no", "L/C 번호", 145, "w"),
           ("open", "개설일", 90, "center"), ("ship", "Latest shipment", 108, "center"),
           ("exp", "Expiry date", 102, "center"), ("exp_dday", "만료", 66, "center"),
           ("rule", "적용규칙", 88, "center"), ("amend", "Amend", 120, "w")]


def row_lc(db, r):
    lead = db.get_int("lc_lead", 7)
    done = db.done_set(r["id"])
    req_due = add_days(r["fca_date"], -lead)
    n = dday(req_due)
    ship, exp = db.lc_rule(r["customer"])
    custom = db.con.execute("SELECT 1 FROM lc_rule WHERE customer=?",
                            (r["customer"] or "",)).fetchone()
    if 4 not in done:
        tag = _tag_by_days(n)
    else:
        tag = "ok" if r["lc_no"] else ""
    return ([r["code"] or "", r["customer"] or "", r["po_no"] or "",
             norm_date(r["fca_date"]), req_due if 4 not in done else "―",
             dday_text(n) if 4 not in done else "완료",
             r["lc_bank"] or "", r["lc_no"] or "", norm_date(r["lc_open_date"]),
             norm_date(r["latest_shipment"]), norm_date(r["expiry_date"]),
             dday_text(dday(r["expiry_date"])),
             "+%d/+%d일%s" % (ship, exp, " ★" if custom else ""),
             r["lc_amend"] or ""], tag)


def sum_lc(rows):
    opened = len([r for r in rows if r[1][7]])
    return "%d건    개설완료 %d건 / 미개설 %d건" % (len(rows), opened, len(rows) - opened)


COLS_SHIP = [("code", "안건번호", 84, "center"), ("customer", "고객사", 118, "w"),
             ("po", "P.O 번호", 116, "w"), ("fca", "FCA", 90, "center"),
             ("ship", "Latest shipment", 110, "center"),
             ("s5", "5. 서류 수령", 98, "center"), ("s6", "6. 서류 발송", 98, "center"),
             ("s7", "7. C/O 발급", 98, "center"), ("exp", "L/C 만료", 92, "center"),
             ("state", "상태", 128, "w"), ("memo", "메모", 220, "w")]


def row_ship(db, r):
    st = db.stages(r["id"])
    done = db.done_set(r["id"])
    if 7 in done:
        state, tag = "서류 완료", "ok"
    elif 6 in done:
        state, tag = "C/O 발급 대기", "soon"
    elif 5 in done:
        state, tag = "고객사 발송 대기", "soon"
    else:
        n = dday(r["latest_shipment"])
        if n is not None and n < 0:
            state, tag = "선적기한 경과", "late"
        elif n is not None and n <= 7:
            state, tag = "선적서류 수령 대기", "soon"
        else:
            state, tag = "선적 전", ""
    memo = " / ".join(x for x in (st.get(5, ("", ""))[1], st.get(6, ("", ""))[1],
                                  st.get(7, ("", ""))[1]) if x)
    return ([r["code"] or "", r["customer"] or "", r["po_no"] or "",
             norm_date(r["fca_date"]), norm_date(r["latest_shipment"]),
             st.get(5, ("", ""))[0] or "―", st.get(6, ("", ""))[0] or "―",
             st.get(7, ("", ""))[0] or "―", norm_date(r["expiry_date"]),
             state, memo], tag)


def sum_ship(rows):
    return "%d건    서류 완료 %d건" % (len(rows), len([r for r in rows if r[2] == "ok"]))


COLS_COMM = [("code", "안건번호", 84, "center"), ("customer", "고객사", 118, "w"),
             ("po", "P.O 번호", 116, "w"), ("amount", "납품금액", 112, "e"),
             ("cur", "통화", 46, "center"), ("rate", "요율%", 54, "center"),
             ("camt", "커미션 금액", 112, "e"), ("inv", "청구서 번호", 190, "w"),
             ("billed", "청구일", 90, "center"), ("passed", "경과", 58, "center"),
             ("recv", "수령일", 90, "center"), ("state", "상태", 96, "center")]


def row_comm(db, r):
    due = db.get_int("comm_due", 30)
    b, v = parse_date(r["comm_billed"]), parse_date(r["comm_received"])
    passed = ""
    if v:
        state, tag = "수령 완료", "ok"
    elif b:
        n = (datetime.date.today() - b).days
        passed = "%d일" % n
        state, tag = ("입금 지연", "late") if n >= due else ("입금 대기", "soon")
    elif 7 in db.done_set(r["id"]):
        state, tag = "청구 필요", "soon"
    else:
        state, tag = "―", ""
    return ([r["code"] or "", r["customer"] or "", r["po_no"] or "",
             money(r["amount"]), r["currency"] or "", r["comm_rate"] or "",
             money(r["comm_amount"]), r["comm_invoice_no"] or "",
             norm_date(r["comm_billed"]), passed, norm_date(r["comm_received"]),
             state], tag)


def sum_comm(rows):
    billed = recv = 0
    for _i, v, _g in rows:
        try:
            amt = int(str(v[6]).replace(",", "") or 0)
        except ValueError:
            amt = 0
        if v[8]:
            billed += amt
        if v[10]:
            recv += amt
    return "%d건    청구 %s    수령 %s    미수 %s" % (
        len(rows), money(billed), money(recv), money(billed - recv))


# (보기키, 라벨, 컬럼, 행 생성 함수, 합계 함수)
VIEWS = [
    ("all",   "전체",            COLS_ALL,   row_all,   sum_all),
    ("order", "발주 · 납기 2~3", COLS_ORDER, row_order, sum_order),
    ("lc",    "신용장 4",        COLS_LC,    row_lc,    sum_lc),
    ("ship",  "선적 · 서류 5~7", COLS_SHIP,  row_ship,  sum_ship),
    ("comm",  "커미션 8",        COLS_COMM,  row_comm,  sum_comm),
]

# --------------------------------------------------------------- 자료방
class ArchivePage(Page):
    """자료방 폴더(00~10 …)를 탐색하고 파일을 바로 여는 화면."""

    def __init__(self, app):
        Page.__init__(self, app)
        self.node_path = {}      # 트리 노드 iid -> 실제 경로
        self.row_path = {}       # 파일목록 iid -> 실제 경로
        self.cur_dir = None
        self._seq = 0
        self._searching = False

        page_title(self.inner, "자료방",
                   "왼쪽에서 분류를 클릭하면 자료가 표시됩니다. "
                   "파일을 더블클릭하면 바로 열립니다.")

        bar = toolbar(self.inner)
        self._btn(bar, "열기", self.open_selected, primary=True)
        self._btn(bar, "탐색기에서 열기", self.open_in_explorer)
        self._btn(bar, "상위 폴더", self.go_up)
        self._btn(bar, "새로고침", self.refresh)
        tk.Label(bar, text="파일 검색", bg=C_BG, fg=C_MUTED,
                 font=FONT_S).pack(side="left", padx=(16, 4))
        self.v_q = tk.StringVar()
        e = ttk.Entry(bar, textvariable=self.v_q, width=24, font=FONT)
        e.pack(side="left")
        e.bind("<Return>", lambda ev: self.do_search())
        self._btn(bar, "찾기", self.do_search)
        self._btn(bar, "지우기", self.clear_search)

        bar2 = toolbar(self.inner)
        self._btn(bar2, "＋ 파일 올리기", self.upload_files, primary=True)
        self._btn(bar2, "새 폴더", self.new_folder)
        self._btn(bar2, "이름 변경", self.rename_item)
        self._btn(bar2, "삭제", self.delete_items)
        tk.Label(bar2, text="※ 파일은 현재 열려 있는 폴더로 복사됩니다. "
                            "원본은 그대로 남습니다.",
                 bg=C_BG, fg=C_MUTED, font=FONT_S).pack(side="left", padx=(14, 0))

        # 분류 칩 (00 ~ 10 …)
        self.chips = tk.Frame(self.inner, bg=C_BG)
        self.chips.pack(fill="x", pady=(0, 10))

        self.v_path = tk.StringVar(value="")
        tk.Label(self.inner, textvariable=self.v_path, bg=C_BG, fg=C_MUTED,
                 font=FONT_S, anchor="w").pack(fill="x", pady=(0, 6))

        pane = ttk.PanedWindow(self.inner, orient="horizontal")
        pane.pack(fill="both", expand=True)

        lf = tk.Frame(pane, bg=C_CARD, highlightthickness=1, highlightbackground=C_LINE)
        pane.add(lf, weight=2)
        self.tree = ttk.Treeview(lf, show="tree", selectmode="browse")
        vs = ttk.Scrollbar(lf, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscroll=vs.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vs.pack(side="right", fill="y")
        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)
        self.tree.bind("<<TreeviewOpen>>", self.on_tree_open)

        rf = tk.Frame(pane, bg=C_BG)
        pane.add(rf, weight=5)
        self.gv = Grid(rf, [("name", "이름", 380, "w"), ("kind", "종류", 90, "center"),
                            ("size", "크기", 90, "e"), ("mtime", "수정한 날짜", 130, "center"),
                            ("where", "위치", 260, "w")],
                       on_open=self.open_row, height=20)
        self.gv.pack(fill="both", expand=True)
        self.v_sum = tk.StringVar()
        tk.Label(rf, textvariable=self.v_sum, bg=C_BG, fg=C_MUTED,
                 font=FONT_S, anchor="w").pack(fill="x", pady=(6, 0))

    def _btn(self, bar, text, cmd, primary=False):
        b = tk.Button(bar, text=text, command=cmd, font=FONT_B if primary else FONT,
                      bg=C_BLUE if primary else "#FFFFFF",
                      fg="#FFFFFF" if primary else C_TEXT, relief="flat", bd=0,
                      padx=13, pady=6, cursor="hand2", highlightthickness=1,
                      highlightbackground=C_LINE)
        b.pack(side="left", padx=(0, 6))
        return b

    # -- 경로 ---------------------------------------------------------------
    def root_path(self):
        return self.db.get_setting("archive_path", DEFAULT_SETTINGS["archive_path"])

    def pick_root(self):
        p = filedialog.askdirectory(parent=self, title="자료방 폴더 선택",
                                    initialdir=os.path.dirname(APP_DIR))
        if p:
            self.db.set_setting("archive_path", os.path.normpath(p))
            self.refresh()

    # -- 화면 ---------------------------------------------------------------
    def refresh(self):
        root = self.root_path()
        for w in self.chips.winfo_children():
            w.destroy()
        for x in self.tree.get_children():
            self.tree.delete(x)
        self.node_path.clear()
        self._seq = 0
        self._searching = False

        if not os.path.isdir(root):
            self.v_path.set("자료방 폴더를 찾을 수 없습니다:  %s" % root)
            tk.Button(self.chips, text="자료방 폴더 선택...", command=self.pick_root,
                      font=FONT_B, relief="flat", bd=0, bg=C_BLUE, fg="#FFFFFF",
                      padx=16, pady=6, cursor="hand2").pack(side="left")
            self.gv.set_rows([])
            self.v_sum.set("")
            return

        rid = self._add_node("", root, os.path.basename(root) or root)
        self.tree.item(rid, open=True)
        self._fill(rid, root)

        for name in self._subdirs(root):
            path = os.path.join(root, name)
            c = tk.Button(self.chips, text=name, font=FONT,
                          command=lambda p=path: self.goto(p),
                          relief="flat", bd=0, bg="#FFFFFF", fg=C_TEXT,
                          activebackground="#E3ECFB", padx=12, pady=5,
                          cursor="hand2", highlightthickness=1,
                          highlightbackground=C_LINE)
            c.pack(side="left", padx=(0, 6), pady=2)

        self.goto(root)

    def _subdirs(self, path):
        try:
            return sorted([n for n in os.listdir(path)
                           if os.path.isdir(os.path.join(path, n))])
        except OSError:
            return []

    def _add_node(self, parent, path, name):
        self._seq += 1
        iid = "n%d" % self._seq
        self.node_path[iid] = path
        self.tree.insert(parent, "end", iid=iid, text="  " + name)
        return iid

    def _fill(self, node, path):
        """자식 폴더를 채우고, 손자가 있으면 더미를 넣어 펼침 표시를 만든다."""
        for name in self._subdirs(path):
            sub = os.path.join(path, name)
            cid = self._add_node(node, sub, name)
            if self._subdirs(sub):
                self.tree.insert(cid, "end", iid=cid + "d", text="…")

    def on_tree_open(self, event=None):
        self.on_tree_open_node(self.tree.focus())

    def on_tree_open_node(self, node):
        """더미 자식('…')을 실제 하위 폴더로 교체 (지연 로딩)."""
        kids = self.tree.get_children(node)
        if len(kids) == 1 and kids[0] not in self.node_path:
            self.tree.delete(kids[0])
            self._fill(node, self.node_path[node])

    def on_tree_select(self, event=None):
        sel = self.tree.selection()
        if sel and sel[0] in self.node_path:
            self.list_dir(self.node_path[sel[0]])

    def goto(self, path):
        """트리에서 해당 경로를 펼쳐 선택하고 목록을 표시."""
        root = self.root_path()
        try:
            rel = os.path.relpath(path, root)
        except ValueError:
            rel = ""
        node = next((i for i, p in self.node_path.items()
                     if os.path.normcase(p) == os.path.normcase(root)), None)
        if node and rel not in (".", ""):
            for part in rel.split(os.sep):
                self.tree.item(node, open=True)
                self.on_tree_open_node(node)
                nxt = None
                for c in self.tree.get_children(node):
                    p = self.node_path.get(c)
                    if p and os.path.basename(p) == part:
                        nxt = c
                        break
                if not nxt:
                    break
                node = nxt
        if node:
            self.tree.item(node, open=True)
            self.tree.selection_set(node)
            self.tree.see(node)
        self.list_dir(path)

    # -- 파일 목록 -----------------------------------------------------------
    def list_dir(self, path):
        self._searching = False
        self.cur_dir = path
        self.row_path.clear()
        root = self.root_path()
        try:
            rel = os.path.relpath(path, os.path.dirname(root))
        except ValueError:
            rel = path
        self.v_path.set("📁  " + rel.replace(os.sep, "   ›   "))

        rows = []
        nd = nf = 0
        total = 0
        try:
            names = sorted(os.listdir(path))
        except OSError as e:
            self.gv.set_rows([])
            self.v_sum.set("폴더를 읽을 수 없습니다: %s" % e)
            return
        dirs = [n for n in names if os.path.isdir(os.path.join(path, n))]
        files = [n for n in names if n not in dirs]
        for name in dirs + files:          # 폴더를 위로
            full = os.path.join(path, name)
            self._seq += 1
            iid = "r%d" % self._seq
            self.row_path[iid] = full
            if name in dirs:
                nd += 1
                try:
                    cnt = len(os.listdir(full))
                except OSError:
                    cnt = 0
                rows.append((iid, ["📁  " + name, "폴더", "%d개" % cnt,
                                   mtime(full), ""], "ok"))
            else:
                nf += 1
                try:
                    sz = os.path.getsize(full)
                except OSError:
                    sz = 0
                total += sz
                rows.append((iid, ["      " + name, file_kind(name), fsize(sz),
                                   mtime(full), ""], ""))
        self.gv.set_rows(rows)
        self.v_sum.set("폴더 %d개 · 파일 %d개    합계 %s" % (nd, nf, fsize(total)))

    def do_search(self):
        q = (self.v_q.get() or "").strip().lower()
        if not q:
            return self.list_dir(self.cur_dir or self.root_path())
        base = self.cur_dir or self.root_path()
        self._searching = True
        self.row_path.clear()
        rows = []
        limit = 500
        for dirpath, dirnames, filenames in os.walk(base):
            for name in sorted(filenames):
                if q not in name.lower():
                    continue
                full = os.path.join(dirpath, name)
                self._seq += 1
                iid = "r%d" % self._seq
                self.row_path[iid] = full
                try:
                    sz = os.path.getsize(full)
                except OSError:
                    sz = 0
                try:
                    where = os.path.relpath(dirpath, base)
                except ValueError:
                    where = dirpath
                rows.append((iid, ["      " + name, file_kind(name), fsize(sz),
                                   mtime(full), where], ""))
                if len(rows) >= limit:
                    break
            if len(rows) >= limit:
                break
        self.gv.set_rows(rows)
        self.v_path.set("🔍  '%s' 검색 결과 — %s 이하" % (
            self.v_q.get().strip(), os.path.basename(base)))
        self.v_sum.set("%d개 찾음%s" % (len(rows), " (상위 %d개만 표시)" % limit
                                     if len(rows) >= limit else ""))

    def clear_search(self):
        self.v_q.set("")
        self.list_dir(self.cur_dir or self.root_path())

    # -- 자료 올리기 / 관리 --------------------------------------------------
    def _target_dir(self):
        """파일을 넣을 대상 폴더. 검색 중이면 검색을 먼저 해제한다."""
        if self._searching:
            self.clear_search()
        d = self.cur_dir
        if not d or not os.path.isdir(d):
            messagebox.showinfo(APP_TITLE, "먼저 왼쪽에서 폴더를 선택하세요.", parent=self)
            return None
        if os.path.normcase(d) == os.path.normcase(self.root_path()):
            if not messagebox.askyesno(
                    APP_TITLE,
                    "자료방 최상위 폴더입니다.\n\n"
                    "보통은 00~10 분류 폴더 안에 넣습니다.\n여기에 그대로 넣을까요?",
                    parent=self):
                return None
        return d

    def _unique(self, folder, name):
        base, ext = os.path.splitext(name)
        i = 2
        while os.path.exists(os.path.join(folder, "%s (%d)%s" % (base, i, ext))):
            i += 1
        return "%s (%d)%s" % (base, i, ext)

    def _ask_conflict(self, name, remain):
        """같은 이름이 있을 때: (선택, 나머지에도 적용 여부)"""
        win = tk.Toplevel(self)
        win.title("같은 이름의 파일이 있습니다")
        win.configure(bg=C_BG)
        win.transient(self)
        win.resizable(False, False)
        win.grab_set()
        result = {"choice": "cancel", "all": False}

        tk.Label(win, text="같은 이름의 파일이 이미 있습니다", font=FONT_H,
                 bg=C_BG, fg=C_TEXT).pack(anchor="w", padx=22, pady=(18, 4))
        tk.Label(win, text=name, font=FONT_B, bg=C_BG, fg=C_BLUE,
                 wraplength=460, justify="left").pack(anchor="w", padx=22)
        tk.Label(win, text="어떻게 할까요?", font=FONT, bg=C_BG,
                 fg=C_MUTED).pack(anchor="w", padx=22, pady=(10, 8))

        v_all = tk.BooleanVar(value=False)
        if remain > 1:
            tk.Checkbutton(win, text="남은 %d개에도 같은 방식 적용" % remain,
                           variable=v_all, bg=C_BG, fg=C_TEXT, font=FONT_S,
                           activebackground=C_BG).pack(anchor="w", padx=20)

        def pick(c):
            result["choice"] = c
            result["all"] = v_all.get()
            win.destroy()

        row = tk.Frame(win, bg=C_BG)
        row.pack(fill="x", padx=20, pady=(12, 18))
        for text, code, primary in (("이름 바꿔 저장", "rename", True),
                                    ("덮어쓰기", "overwrite", False),
                                    ("건너뛰기", "skip", False),
                                    ("취소", "cancel", False)):
            tk.Button(row, text=text, command=lambda c=code: pick(c),
                      font=FONT_B if primary else FONT,
                      bg=C_BLUE if primary else "#FFFFFF",
                      fg="#FFFFFF" if primary else C_TEXT, relief="flat", bd=0,
                      padx=14, pady=6, cursor="hand2", highlightthickness=1,
                      highlightbackground=C_LINE).pack(side="left", padx=(0, 8))
        win.update_idletasks()
        win.geometry("+%d+%d" % (self.winfo_rootx() + 240, self.winfo_rooty() + 200))
        self.wait_window(win)
        return result["choice"], result["all"]

    def upload_files(self):
        dest = self._target_dir()
        if not dest:
            return
        paths = filedialog.askopenfilenames(
            parent=self, title="자료방에 올릴 파일 선택  →  %s" % os.path.basename(dest),
            filetypes=[("모든 파일", "*.*"),
                       ("문서", "*.xlsx;*.xls;*.docx;*.doc;*.pptx;*.ppt;*.pdf;*.hwp"),
                       ("이미지", "*.png;*.jpg;*.jpeg;*.gif;*.bmp"),
                       ("압축", "*.zip;*.7z;*.rar")])
        if not paths:
            return

        done, skipped, failed = [], 0, []
        rule = None
        for i, src in enumerate(paths):
            name = os.path.basename(src)
            dst = os.path.join(dest, name)
            if os.path.exists(dst):
                choice = rule
                if choice is None:
                    choice, apply_all = self._ask_conflict(name, len(paths) - i)
                    if apply_all:
                        rule = choice
                if choice == "cancel":
                    break
                if choice == "skip":
                    skipped += 1
                    continue
                if choice == "rename":
                    dst = os.path.join(dest, self._unique(dest, name))
            try:
                shutil.copy2(src, dst)
                done.append(os.path.basename(dst))
            except Exception as e:
                failed.append("%s — %s" % (name, e))

        self.list_dir(dest)
        for iid, path in self.row_path.items():
            if os.path.basename(path) in done:
                self.gv.tv.selection_add(iid)
                self.gv.tv.see(iid)

        msg = "%d개 파일을 올렸습니다.\n\n대상 폴더\n%s" % (len(done), dest)
        if skipped:
            msg += "\n\n건너뜀 : %d개" % skipped
        if failed:
            msg += "\n\n실패 : %d개\n%s" % (len(failed), "\n".join(failed[:5]))
            messagebox.showwarning(APP_TITLE, msg, parent=self)
        elif done:
            messagebox.showinfo(APP_TITLE, msg, parent=self)
        self.app.set_status("자료방 — %d개 올림 (%s)" % (len(done), os.path.basename(dest)))

    def new_folder(self):
        dest = self._target_dir()
        if not dest:
            return
        win = tk.Toplevel(self)
        win.title("새 폴더")
        win.configure(bg=C_BG)
        win.transient(self)
        win.resizable(False, False)
        win.grab_set()
        tk.Label(win, text="새 폴더 이름", font=FONT_B, bg=C_BG,
                 fg=C_TEXT).pack(anchor="w", padx=20, pady=(18, 2))
        tk.Label(win, text=dest, font=FONT_S, bg=C_BG, fg=C_MUTED,
                 wraplength=420, justify="left").pack(anchor="w", padx=20)
        v = tk.StringVar()
        ent = ttk.Entry(win, textvariable=v, width=46, font=FONT)
        ent.pack(padx=20, pady=(10, 4))
        ent.focus_set()

        def ok():
            name = v.get().strip()
            if not name:
                return
            if re.search(r'[\\/:*?"<>|]', name):
                messagebox.showerror(APP_TITLE, '폴더 이름에 \\ / : * ? " < > | 는 쓸 수 없습니다.',
                                     parent=win)
                return
            path = os.path.join(dest, name)
            if os.path.exists(path):
                messagebox.showerror(APP_TITLE, "이미 같은 이름의 폴더가 있습니다.", parent=win)
                return
            try:
                os.makedirs(path)
            except Exception as e:
                messagebox.showerror(APP_TITLE, "폴더를 만들 수 없습니다.\n\n%s" % e, parent=win)
                return
            win.destroy()
            self.refresh()
            self.goto(path)
            self.app.set_status("새 폴더 — %s" % name)

        ent.bind("<Return>", lambda e: ok())
        row = tk.Frame(win, bg=C_BG)
        row.pack(fill="x", padx=20, pady=(6, 18))
        tk.Button(row, text="만들기", command=ok, font=FONT_B, bg=C_BLUE, fg="#FFFFFF",
                  relief="flat", bd=0, padx=18, pady=6, cursor="hand2").pack(side="right")
        tk.Button(row, text="취소", command=win.destroy, font=FONT, bg="#FFFFFF",
                  fg=C_TEXT, relief="flat", bd=0, padx=18, pady=6, cursor="hand2",
                  highlightthickness=1, highlightbackground=C_LINE
                  ).pack(side="right", padx=(0, 8))
        win.update_idletasks()
        win.geometry("+%d+%d" % (self.winfo_rootx() + 260, self.winfo_rooty() + 220))

    def rename_item(self):
        sel = self.gv.selection()
        if not sel:
            messagebox.showinfo(APP_TITLE, "이름을 바꿀 항목을 선택하세요.", parent=self)
            return
        old = self.row_path.get(sel[0])
        if not old:
            return
        folder, name = os.path.split(old)
        base, ext = os.path.splitext(name)

        win = tk.Toplevel(self)
        win.title("이름 변경")
        win.configure(bg=C_BG)
        win.transient(self)
        win.resizable(False, False)
        win.grab_set()
        tk.Label(win, text="새 이름", font=FONT_B, bg=C_BG,
                 fg=C_TEXT).pack(anchor="w", padx=20, pady=(18, 6))
        v = tk.StringVar(value=name)
        ent = ttk.Entry(win, textvariable=v, width=54, font=FONT)
        ent.pack(padx=20)
        ent.focus_set()
        ent.selection_range(0, len(base))     # 확장자는 빼고 선택

        def ok():
            new = v.get().strip()
            if not new or new == name:
                win.destroy()
                return
            if re.search(r'[\\/:*?"<>|]', new):
                messagebox.showerror(APP_TITLE, '이름에 \\ / : * ? " < > | 는 쓸 수 없습니다.',
                                     parent=win)
                return
            dst = os.path.join(folder, new)
            if os.path.exists(dst):
                messagebox.showerror(APP_TITLE, "이미 같은 이름이 있습니다.", parent=win)
                return
            try:
                os.rename(old, dst)
            except Exception as e:
                messagebox.showerror(APP_TITLE, "이름을 바꿀 수 없습니다.\n\n%s" % e, parent=win)
                return
            win.destroy()
            if os.path.isdir(dst):
                self.refresh()
                self.goto(folder)
            else:
                self.list_dir(folder)
            self.app.set_status("이름 변경 — %s" % new)

        ent.bind("<Return>", lambda e: ok())
        row = tk.Frame(win, bg=C_BG)
        row.pack(fill="x", padx=20, pady=(12, 18))
        tk.Button(row, text="변경", command=ok, font=FONT_B, bg=C_BLUE, fg="#FFFFFF",
                  relief="flat", bd=0, padx=18, pady=6, cursor="hand2").pack(side="right")
        tk.Button(row, text="취소", command=win.destroy, font=FONT, bg="#FFFFFF",
                  fg=C_TEXT, relief="flat", bd=0, padx=18, pady=6, cursor="hand2",
                  highlightthickness=1, highlightbackground=C_LINE
                  ).pack(side="right", padx=(0, 8))
        win.update_idletasks()
        win.geometry("+%d+%d" % (self.winfo_rootx() + 240, self.winfo_rooty() + 220))

    def delete_items(self):
        sel = self.gv.selection()
        if not sel:
            messagebox.showinfo(APP_TITLE, "삭제할 항목을 선택하세요.", parent=self)
            return
        paths = [self.row_path[i] for i in sel if i in self.row_path]
        names = [os.path.basename(p) for p in paths]
        has_dir = any(os.path.isdir(p) for p in paths)
        if not messagebox.askyesno(
                APP_TITLE,
                "%d개 항목을 휴지통으로 보냅니다.%s\n\n%s%s\n\n진행할까요?"
                % (len(paths), "\n(폴더는 안의 내용까지 함께 삭제됩니다)" if has_dir else "",
                   "\n".join(names[:8]),
                   "\n… 외 %d개" % (len(names) - 8) if len(names) > 8 else ""),
                parent=self):
            return

        folder = os.path.dirname(paths[0])
        ok, fail = recycle(paths)
        if fail:
            if messagebox.askyesno(
                    APP_TITLE,
                    "휴지통으로 보내지 못한 항목이 %d개 있습니다.\n\n"
                    "완전히 삭제할까요?  (되돌릴 수 없습니다)" % len(fail), parent=self):
                for p in list(fail):
                    try:
                        if os.path.isdir(p):
                            shutil.rmtree(p)
                        else:
                            os.remove(p)
                        ok += 1
                        fail.remove(p)
                    except Exception:
                        pass
        if has_dir:
            self.refresh()
            self.goto(folder)
        else:
            self.list_dir(folder)
        self.app.set_status("삭제 %d개%s" % (ok, " (실패 %d개)" % len(fail) if fail else ""))

    # -- 열기 ---------------------------------------------------------------
    def open_row(self, iid):
        path = self.row_path.get(iid)
        if not path:
            return
        if os.path.isdir(path):
            self.goto(path)
        else:
            open_path(path, self)
            self.app.set_status("열기 — %s" % os.path.basename(path))

    def open_selected(self):
        sel = self.gv.selection()
        if sel:
            self.open_row(sel[0])
        elif self.cur_dir:
            open_path(self.cur_dir, self)

    def open_in_explorer(self):
        sel = self.gv.selection()
        target = self.row_path.get(sel[0]) if sel else self.cur_dir
        if not target:
            return
        if not os.path.isdir(target):
            target = os.path.dirname(target)
        open_path(target, self)

    def go_up(self):
        if not self.cur_dir:
            return
        root = self.root_path()
        if os.path.normcase(self.cur_dir) == os.path.normcase(root):
            return
        self.goto(os.path.dirname(self.cur_dir))


# ------------------------------------------------------------- 기준정보
class MasterPage(Page):
    def __init__(self, app, parent=None):
        Page.__init__(self, app, parent)
        page_title(self.inner, "설정 · 기준정보",
                   "목록 선택값, 신용장 기한 규칙, 알림 기준, 자료방 경로를 관리합니다.")

        wrap = tk.Frame(self.inner, bg=C_BG)
        wrap.pack(fill="both", expand=True)
        wrap.columnconfigure(0, weight=2)
        wrap.columnconfigure(1, weight=3)
        wrap.rowconfigure(0, weight=1)

        # --- 코드 목록
        box = tk.LabelFrame(wrap, text=" 선택 목록 ", bg=C_BG, fg=C_TEXT, font=FONT_B, bd=0)
        box.grid(row=0, column=0, sticky="nsew", padx=(0, 14))
        row = tk.Frame(box, bg=C_BG)
        row.pack(fill="x", pady=(8, 6))
        tk.Label(row, text="분류", bg=C_BG, fg=C_MUTED, font=FONT_S).pack(side="left")
        self.v_cat = tk.StringVar(value=MASTER_CATS[0])
        cb = ttk.Combobox(row, textvariable=self.v_cat, values=MASTER_CATS, width=12,
                          state="readonly", font=FONT)
        cb.pack(side="left", padx=(6, 0))
        cb.bind("<<ComboboxSelected>>", lambda e: self.reload_codes())

        self.lb = tk.Listbox(box, font=FONT, relief="flat", highlightthickness=1,
                             highlightbackground=C_LINE, activestyle="none")
        self.lb.pack(fill="both", expand=True, pady=(0, 6))

        add = tk.Frame(box, bg=C_BG)
        add.pack(fill="x")
        self.v_new = tk.StringVar()
        ent = ttk.Entry(add, textvariable=self.v_new, font=FONT)
        ent.pack(side="left", fill="x", expand=True)
        ent.bind("<Return>", lambda e: self.add_code())
        tk.Button(add, text="추가", command=self.add_code, font=FONT, relief="flat",
                  bd=0, bg=C_BLUE, fg="#FFFFFF", padx=14, pady=4, cursor="hand2"
                  ).pack(side="left", padx=(6, 0))
        tk.Button(add, text="삭제", command=self.del_code, font=FONT, relief="flat",
                  bd=0, bg="#FFFFFF", fg=C_TEXT, padx=14, pady=4, cursor="hand2",
                  highlightthickness=1, highlightbackground=C_LINE
                  ).pack(side="left", padx=(6, 0))

        # --- 오른쪽
        right = tk.Frame(wrap, bg=C_BG)
        right.grid(row=0, column=1, sticky="nsew")
        right.rowconfigure(0, weight=1)
        right.columnconfigure(0, weight=1)

        rule = tk.LabelFrame(right, text=" 신용장 기한 규칙 (고객사별) ", bg=C_BG,
                             fg=C_TEXT, font=FONT_B, bd=0)
        rule.grid(row=0, column=0, sticky="nsew", pady=(0, 14))
        tk.Label(rule, bg=C_BG, fg=C_MUTED, font=FONT_S, justify="left",
                 text="여기에 없는 고객사는 아래 [기본값] 이 적용됩니다.  "
                      "기본 +2주/+4주. 고객사별로 다르게 둘 수 있습니다"
                 ).pack(anchor="w", pady=(6, 6))
        self.g_rule = Grid(rule, [("customer", "고객사", 160, "w"),
                                  ("ship", "Latest shipment", 130, "center"),
                                  ("exp", "Expiry date", 130, "center")],
                           height=6, selectmode="browse")
        self.g_rule.pack(fill="both", expand=True)

        rf = tk.Frame(rule, bg=C_BG)
        rf.pack(fill="x", pady=(8, 4))
        tk.Label(rf, text="고객사", bg=C_BG, fg=C_MUTED, font=FONT_S).pack(side="left")
        self.v_rc = tk.StringVar()
        self.cb_rc = ttk.Combobox(rf, textvariable=self.v_rc, width=16, font=FONT)
        self.cb_rc.pack(side="left", padx=(6, 12))
        tk.Label(rf, text="선적 +", bg=C_BG, fg=C_MUTED, font=FONT_S).pack(side="left")
        self.v_rs = tk.StringVar(value="7")
        ttk.Entry(rf, textvariable=self.v_rs, width=5, font=FONT).pack(side="left", padx=(4, 2))
        tk.Label(rf, text="일    만료 +", bg=C_BG, fg=C_MUTED, font=FONT_S).pack(side="left")
        self.v_re = tk.StringVar(value="14")
        ttk.Entry(rf, textvariable=self.v_re, width=5, font=FONT).pack(side="left", padx=(4, 2))
        tk.Label(rf, text="일", bg=C_BG, fg=C_MUTED, font=FONT_S).pack(side="left")
        tk.Button(rf, text="저장", command=self.save_rule, font=FONT, relief="flat",
                  bd=0, bg=C_BLUE, fg="#FFFFFF", padx=14, pady=3, cursor="hand2"
                  ).pack(side="left", padx=(12, 0))
        tk.Button(rf, text="삭제", command=self.del_rule, font=FONT, relief="flat",
                  bd=0, bg="#FFFFFF", fg=C_TEXT, padx=14, pady=3, cursor="hand2",
                  highlightthickness=1, highlightbackground=C_LINE
                  ).pack(side="left", padx=(6, 0))

        cfg = tk.LabelFrame(right, text=" 기본값 · 알림 기준 ", bg=C_BG, fg=C_TEXT,
                            font=FONT_B, bd=0)
        cfg.grid(row=1, column=0, sticky="ew")
        g = tk.Frame(cfg, bg=C_BG)
        g.pack(fill="x", pady=(8, 8))
        self.sv = {}
        spec = [("company", "회사명", 22), ("user", "담당자명", 22),
                ("ship_days", "기본 Latest shipment (FCA + n일)", 6),
                ("expiry_days", "기본 Expiry date (FCA + n일)", 6),
                ("lc_lead", "L/C 개설요청 기한 (FCA − n일)", 6),
                ("comm_due", "커미션 입금 기준 (청구 후 n일)", 6),
                ("horizon", "알림 표시 범위 (n일 앞까지)", 6)]
        for i, (key, label, width) in enumerate(spec):
            r, c = divmod(i, 2)
            cell = tk.Frame(g, bg=C_BG)
            cell.grid(row=r, column=c, sticky="w", padx=(0, 24), pady=4)
            tk.Label(cell, text=label, bg=C_BG, fg=C_TEXT, font=FONT,
                     width=28, anchor="w").pack(side="left")
            v = tk.StringVar()
            self.sv[key] = v
            ttk.Entry(cell, textvariable=v, width=width, font=FONT).pack(side="left")

        prow = tk.Frame(cfg, bg=C_BG)
        prow.pack(fill="x", pady=(6, 2))
        tk.Label(prow, text="자료방 폴더", bg=C_BG, fg=C_TEXT, font=FONT,
                 width=28, anchor="w").pack(side="left")
        self.sv["archive_path"] = tk.StringVar()
        ttk.Entry(prow, textvariable=self.sv["archive_path"], font=FONT
                  ).pack(side="left", fill="x", expand=True)
        tk.Button(prow, text="폴더 선택", command=self.pick_archive, font=FONT,
                  relief="flat", bd=0, bg="#FFFFFF", fg=C_TEXT, padx=12, pady=3,
                  cursor="hand2", highlightthickness=1, highlightbackground=C_LINE
                  ).pack(side="left", padx=(6, 0))

        tk.Button(cfg, text="설정 저장", command=self.save_settings, font=FONT_B,
                  relief="flat", bd=0, bg=C_BLUE, fg="#FFFFFF", padx=18, pady=5,
                  cursor="hand2").pack(anchor="e", pady=(4, 10))

    def pick_archive(self):
        cur = self.sv["archive_path"].get().strip()
        p = filedialog.askdirectory(
            parent=self, title="자료방 폴더 선택",
            initialdir=cur if os.path.isdir(cur) else os.path.dirname(APP_DIR))
        if p:
            self.sv["archive_path"].set(os.path.normpath(p))

    def refresh(self):
        self.reload_codes()
        self.reload_rules()
        for k, v in self.sv.items():
            v.set(self.db.get_setting(k, DEFAULT_SETTINGS.get(k, "")))

    def reload_codes(self):
        self.lb.delete(0, "end")
        for v in self.db.codes(self.v_cat.get()):
            self.lb.insert("end", v)

    def reload_rules(self):
        self.cb_rc.configure(values=self.db.codes("고객사"))
        self.g_rule.set_rows([
            (r["customer"], [r["customer"], "+%d일 (%d주)" % (r["ship_days"], r["ship_days"] // 7),
                             "+%d일 (%d주)" % (r["expiry_days"], r["expiry_days"] // 7)], "")
            for r in self.db.lc_rules()])

    def add_code(self):
        v = self.v_new.get().strip()
        if not v:
            return
        self.db.add_code(self.v_cat.get(), v)
        self.v_new.set("")
        self.reload_codes()
        self.reload_rules()
        self.app.refresh_others("master")

    def del_code(self):
        sel = self.lb.curselection()
        if not sel:
            return
        v = self.lb.get(sel[0])
        if messagebox.askyesno(APP_TITLE, "'%s' 을(를) 목록에서 삭제할까요?\n"
                                          "(이미 입력된 안건의 값은 그대로 남습니다)" % v,
                               parent=self):
            self.db.del_code(self.v_cat.get(), v)
            self.reload_codes()
            self.reload_rules()
            self.app.refresh_others("master")

    def save_rule(self):
        c = self.v_rc.get().strip()
        if not c:
            messagebox.showinfo(APP_TITLE, "고객사를 입력하거나 선택하세요.", parent=self)
            return
        try:
            s, e = int(self.v_rs.get()), int(self.v_re.get())
        except ValueError:
            messagebox.showerror(APP_TITLE, "가산일수는 숫자로 입력하세요.", parent=self)
            return
        self.db.set_lc_rule(c, s, e)
        self.reload_rules()
        self.app.set_status("신용장 규칙 저장 — %s (+%d일 / +%d일)" % (c, s, e))

    def del_rule(self):
        sel = self.g_rule.selection()
        if not sel:
            return
        self.db.del_lc_rule(sel[0])
        self.reload_rules()
        self.app.set_status("규칙 삭제 — %s (기본값 적용)" % sel[0])

    def save_settings(self):
        for k, v in self.sv.items():
            val = v.get().strip()
            if k in NUMERIC_SETTINGS:
                try:
                    int(val)
                except ValueError:
                    messagebox.showerror(APP_TITLE,
                                         "숫자로 입력해야 하는 항목이 있습니다.", parent=self)
                    return
            self.db.set_setting(k, val)
        self.app.refresh_header()
        self.app.refresh_others("master")
        self.app.set_status("설정을 저장했습니다.")




# --------------------------------------------------------- 업무 플로우
FLOW_COLORS = {0: C_BLUE, 1: C_BLUE, 2: "#5B4FCF", 3: "#5B4FCF", 4: C_AMBER,
               5: C_GREEN, 6: C_GREEN, 7: C_GREEN, 8: "#0E7C86"}
PALETTE = [("자동(단계색)", ""), ("파랑", C_BLUE), ("보라", "#5B4FCF"),
           ("주황", C_AMBER), ("초록", C_GREEN), ("청록", "#0E7C86"),
           ("빨강", C_RED), ("회색", C_GRAY)]
SHAPES = [("사각형 (작업)", "box"), ("타원 (시작·종료)", "oval"),
          ("마름모 (판단)", "diamond")]
NODE_W, NODE_H = 212, 86
SNAP = 5


def flow_color(n):
    c = (n["color"] or "").strip()
    if c:
        return c
    sn = n["stage_no"]
    if sn is None or sn < 0:
        return C_GRAY
    return FLOW_COLORS.get(int(sn), C_GRAY)


class FlowEditor(tk.Toplevel):
    """플로우 항목 추가 / 수정 창."""

    def __init__(self, page, node=None):
        tk.Toplevel.__init__(self, page)
        self.page = page
        self.db = page.db
        self.node = node
        self.title("플로우 항목 " + ("수정" if node else "추가"))
        self.configure(bg=C_BG)
        self.transient(page.winfo_toplevel())
        self.geometry("640x660")
        self.grab_set()

        body = tk.Frame(self, bg=C_BG)
        body.pack(fill="both", expand=True, padx=22, pady=(18, 8))
        body.columnconfigure(1, weight=1)

        self.v = {}
        row = 0
        for key, label in (("title", "항목 이름"), ("store", "저장소 (안내문)")):
            tk.Label(body, text=label, font=FONT_B, bg=C_BG, fg=C_TEXT
                     ).grid(row=row, column=0, sticky="w", pady=(8, 2))
            self.v[key] = tk.StringVar(value=(node[key] if node else ""))
            ttk.Entry(body, textvariable=self.v[key], font=FONT
                      ).grid(row=row, column=1, sticky="ew", padx=(12, 0), pady=(8, 2))
            row += 1

        tk.Label(body, text="설명", font=FONT_B, bg=C_BG, fg=C_TEXT
                 ).grid(row=row, column=0, sticky="nw", pady=(10, 2))
        self.t_desc = tk.Text(body, font=FONT, height=6, wrap="word", relief="flat",
                              bg=C_CARD, padx=8, pady=6, highlightthickness=1,
                              highlightbackground=C_LINE)
        self.t_desc.grid(row=row, column=1, sticky="ew", padx=(12, 0), pady=(10, 2))
        if node:
            self.t_desc.insert("1.0", node["descr"] or "")
        row += 1

        tk.Label(body, text="체크포인트", font=FONT_B, bg=C_BG, fg=C_TEXT
                 ).grid(row=row, column=0, sticky="nw", pady=(10, 2))
        self.t_tip = tk.Text(body, font=FONT, height=3, wrap="word", relief="flat",
                             bg=C_CARD, padx=8, pady=6, highlightthickness=1,
                             highlightbackground=C_LINE)
        self.t_tip.grid(row=row, column=1, sticky="ew", padx=(12, 0), pady=(10, 2))
        if node:
            self.t_tip.insert("1.0", node["tip"] or "")
        row += 1

        # 도형 · 색
        tk.Label(body, text="도형 / 색", font=FONT_B, bg=C_BG, fg=C_TEXT
                 ).grid(row=row, column=0, sticky="w", pady=(12, 2))
        sc = tk.Frame(body, bg=C_BG)
        sc.grid(row=row, column=1, sticky="ew", padx=(12, 0), pady=(12, 2))
        cur_shape = (node["shape"] if node and node["shape"] else "box")
        self.v["shape"] = tk.StringVar(
            value=dict((v, k) for k, v in SHAPES).get(cur_shape, SHAPES[0][0]))
        ttk.Combobox(sc, textvariable=self.v["shape"], values=[s[0] for s in SHAPES],
                     state="readonly", width=17, font=FONT).pack(side="left")
        cur_color = (node["color"] if node and node["color"] else "")
        self.v["color"] = tk.StringVar(
            value=dict((v, k) for k, v in PALETTE).get(cur_color, PALETTE[0][0]))
        ttk.Combobox(sc, textvariable=self.v["color"], values=[p[0] for p in PALETTE],
                     state="readonly", width=15, font=FONT).pack(side="left", padx=(8, 0))
        row += 1

        tk.Label(body, text="자료방 폴더", font=FONT_B, bg=C_BG, fg=C_TEXT
                 ).grid(row=row, column=0, sticky="w", pady=(10, 2))
        fr = tk.Frame(body, bg=C_BG)
        fr.grid(row=row, column=1, sticky="ew", padx=(12, 0), pady=(10, 2))
        self.v["folder"] = tk.StringVar(value=(node["folder"] if node else ""))
        ttk.Entry(fr, textvariable=self.v["folder"], font=FONT
                  ).pack(side="left", fill="x", expand=True)
        tk.Button(fr, text="찾아보기", command=self.pick_folder, font=FONT_S,
                  relief="flat", bd=0, bg="#FFFFFF", fg=C_TEXT, padx=10, pady=3,
                  cursor="hand2", highlightthickness=1, highlightbackground=C_LINE
                  ).pack(side="left", padx=(6, 0))
        row += 1
        tk.Label(body, text="비워두면 연결 단계의 기본 폴더를 사용합니다.",
                 font=FONT_S, bg=C_BG, fg=C_MUTED
                 ).grid(row=row, column=1, sticky="w", padx=(12, 0))
        row += 1

        tk.Label(body, text="연결 단계", font=FONT_B, bg=C_BG, fg=C_TEXT
                 ).grid(row=row, column=0, sticky="w", pady=(12, 2))
        opts = ["(연결 안 함)"] + ["%d. %s" % (n, STAGE_NAME[n]) for n in range(STAGE_COUNT)]
        cur = "(연결 안 함)"
        if node and node["stage_no"] is not None and 0 <= node["stage_no"] < STAGE_COUNT:
            cur = opts[node["stage_no"] + 1]
        self.v["stage"] = tk.StringVar(value=cur)
        ttk.Combobox(body, textvariable=self.v["stage"], values=opts, state="readonly",
                     font=FONT).grid(row=row, column=1, sticky="ew", padx=(12, 0),
                                     pady=(12, 2))
        row += 1
        tk.Label(body, text="연결하면 그 단계의 진행중 안건이 이 항목에 함께 표시됩니다.",
                 font=FONT_S, bg=C_BG, fg=C_MUTED
                 ).grid(row=row, column=1, sticky="w", padx=(12, 0))

        bar = tk.Frame(self, bg=C_BG)
        bar.pack(fill="x", padx=22, pady=(6, 18))
        tk.Button(bar, text="저장", command=self.save, font=FONT_B, relief="flat",
                  bd=0, bg=C_BLUE, fg="#FFFFFF", padx=22, pady=7,
                  cursor="hand2").pack(side="right")
        tk.Button(bar, text="취소", command=self.destroy, font=FONT, relief="flat",
                  bd=0, bg="#FFFFFF", fg=C_TEXT, padx=20, pady=7, cursor="hand2",
                  highlightthickness=1, highlightbackground=C_LINE
                  ).pack(side="right", padx=(0, 8))

    def pick_folder(self):
        root = self.page.archive_root()
        cur = self.v["folder"].get().strip()
        start = os.path.join(root, cur) if cur else root
        p = filedialog.askdirectory(parent=self, title="자료방 폴더 선택",
                                    initialdir=start if os.path.isdir(start) else root)
        if not p:
            return
        p = os.path.normpath(p)
        try:
            rel = os.path.relpath(p, root)
        except ValueError:
            rel = p
        self.v["folder"].set(p if rel.startswith("..") else rel)

    def save(self):
        title = self.v["title"].get().strip()
        if not title:
            messagebox.showinfo(APP_TITLE, "항목 이름을 입력하세요.", parent=self)
            return
        s = self.v["stage"].get()
        stage_no = -1
        if s != "(연결 안 함)":
            try:
                stage_no = int(s.split(".")[0])
            except ValueError:
                stage_no = -1
        data = {"title": title, "store": self.v["store"].get().strip(),
                "descr": self.t_desc.get("1.0", "end").strip(),
                "tip": self.t_tip.get("1.0", "end").strip(),
                "folder": self.v["folder"].get().strip(), "stage_no": stage_no,
                "shape": dict(SHAPES).get(self.v["shape"].get(), "box"),
                "color": dict(PALETTE).get(self.v["color"].get(), "")}
        if self.node:
            self.db.update_flow(self.node["id"], data)
            nid = self.node["id"]
        else:
            data["x"], data["y"] = self.page.free_spot()
            nid = self.db.add_flow(data)
        self.destroy()
        self.page.reload()
        self.page.select("node", nid)
        self.page.app.set_status("플로우 항목 저장 — %s" % title)


class FlowPage(Page):
    """자유롭게 그리는 업무 플로우차트.
       · 박스를 끌어서 원하는 위치에 배치
       · [연결 그리기] 로 항목끼리 화살표 연결
       · 항목을 클릭하면 관련 자료·안건 표시"""

    def __init__(self, app):
        Page.__init__(self, app)
        self.sel_kind = None          # "node" | "edge" | None
        self.sel_id = None
        self.item_map = {}            # canvas item -> (kind, id)
        self.node_tag = {}            # node id -> canvas tag
        self.npos = {}                # node id -> [x, y]
        self.nodes = []
        self.edges = []
        self.file_rows = {}
        self.connect_mode = False
        self.connect_from = None
        self._drag = None
        self.scale = 1.0
        self._fitted = False
        self.detail_on = True

        page_title(self.inner, "업무 플로우",
                   "박스를 끌어 자유롭게 배치하고, [연결 그리기] 로 화살표를 이으세요. "
                   "[화면에 맞추기] 로 전체를 한눈에 볼 수 있습니다.")

        bar = toolbar(self.inner)
        self._btn(bar, "＋ 항목 추가", self.add_node, primary=True)
        self._btn(bar, "수정", self.edit_node)
        self._btn(bar, "삭제", self.delete_sel)
        tk.Frame(bar, bg=C_LINE, width=1).pack(side="left", fill="y", padx=10, pady=3)
        self.b_conn = self._btn(bar, "🔗 연결 그리기", self.toggle_connect)
        self._btn(bar, "자동 정렬", self.auto_layout)
        self._btn(bar, "기본값 복원", self.reset)
        self.v_hint = tk.StringVar(value="")
        tk.Label(bar, textvariable=self.v_hint, bg=C_BG, fg=C_BLUE,
                 font=FONT_SB).pack(side="left", padx=(14, 0))

        # --- 보기 (확대·축소 · 한눈에 보기)
        self.b_detail = self._btn(bar, "상세 숨기기", self.toggle_detail)
        self.b_detail.pack_forget()
        self.b_detail.pack(side="right", padx=(6, 0))
        self._btn(bar, "＋", lambda: self.zoom(1.25)).pack_configure(side="right", padx=(2, 0))
        self.v_zoom = tk.StringVar(value="100%")
        tk.Label(bar, textvariable=self.v_zoom, bg=C_BG, fg=C_TEXT, font=FONT_SB,
                 width=5).pack(side="right")
        self._btn(bar, "－", lambda: self.zoom(0.8)).pack_configure(side="right", padx=(6, 2))
        self._btn(bar, "🔍 화면에 맞추기", self.fit).pack_configure(side="right", padx=(6, 8))

        pane = ttk.PanedWindow(self.inner, orient="horizontal")
        pane.pack(fill="both", expand=True)
        self.pane = pane

        left = tk.Frame(pane, bg=C_CARD, highlightthickness=1,
                        highlightbackground=C_LINE)
        pane.add(left, weight=5)
        self.cv = tk.Canvas(left, bg="#FAFBFD", highlightthickness=0)
        vs = ttk.Scrollbar(left, orient="vertical", command=self.cv.yview)
        hs = ttk.Scrollbar(left, orient="horizontal", command=self.cv.xview)
        self.cv.configure(yscrollcommand=vs.set, xscrollcommand=hs.set)
        self.cv.grid(row=0, column=0, sticky="nsew")
        vs.grid(row=0, column=1, sticky="ns")
        hs.grid(row=1, column=0, sticky="ew")
        left.rowconfigure(0, weight=1)
        left.columnconfigure(0, weight=1)

        self.cv.bind("<Button-1>", self.on_press)
        self.cv.bind("<B1-Motion>", self.on_drag)
        self.cv.bind("<ButtonRelease-1>", self.on_release)
        self.cv.bind("<Double-1>", self.on_double)
        self.cv.bind("<Escape>", lambda e: self.cancel_connect())
        self.cv.bind("<Configure>", self.on_canvas_resize)
        self.cv.bind("<Enter>", lambda e: self.cv.bind_all("<MouseWheel>", self.on_wheel))
        self.cv.bind("<Leave>", lambda e: self.cv.unbind_all("<MouseWheel>"))

        right = tk.Frame(pane, bg=C_BG)
        pane.add(right, weight=4)
        self.right = right

        self.v_ntitle = tk.StringVar(value="항목을 클릭하세요")
        tk.Label(right, textvariable=self.v_ntitle, font=FONT_H, bg=C_BG, fg=C_TEXT,
                 anchor="w", wraplength=380, justify="left").pack(fill="x", pady=(0, 6))

        self.txt = tk.Text(right, font=FONT, height=7, wrap="word", relief="flat",
                           bg=C_CARD, padx=12, pady=10, highlightthickness=1,
                           highlightbackground=C_LINE, cursor="arrow")
        self.txt.pack(fill="x")
        self.txt.tag_configure("tip", foreground=C_AMBER, font=FONT_S)
        self.txt.tag_configure("store", foreground=C_MUTED, font=FONT_S)
        self.txt.configure(state="disabled")

        frow = tk.Frame(right, bg=C_BG)
        frow.pack(fill="x", pady=(8, 4))
        self.v_folder = tk.StringVar(value="")
        tk.Label(frow, textvariable=self.v_folder, font=FONT_S, bg=C_BG, fg=C_MUTED,
                 anchor="w").pack(side="left", fill="x", expand=True)
        tk.Button(frow, text="📁 자료방에서 열기", command=self.open_in_archive,
                  font=FONT_S, relief="flat", bd=0, bg="#FFFFFF", fg=C_BLUE,
                  padx=10, pady=3, cursor="hand2", highlightthickness=1,
                  highlightbackground=C_LINE).pack(side="left", padx=(6, 0))

        tk.Label(right, text="관련 자료", font=FONT_B, bg=C_BG, fg=C_TEXT,
                 anchor="w").pack(fill="x", pady=(10, 2))
        self.g_files = Grid(right, [("name", "이름", 240, "w"),
                                    ("kind", "종류", 70, "center"),
                                    ("size", "크기", 72, "e"),
                                    ("mtime", "수정한 날짜", 118, "center")],
                            on_open=self.open_file, height=8)
        self.g_files.pack(fill="both", expand=True)

        tk.Label(right, text="이 단계 진행중 안건", font=FONT_B, bg=C_BG, fg=C_TEXT,
                 anchor="w").pack(fill="x", pady=(10, 2))
        self.g_deals = Grid(right, [("code", "안건번호", 82, "center"),
                                    ("customer", "고객사", 110, "w"),
                                    ("model", "모델 / 품명", 160, "w"),
                                    ("next", "다음 기한", 130, "w")],
                            on_open=lambda iid: self.app.open_deal(int(iid)), height=6)
        self.g_deals.pack(fill="both", expand=True)

    def _btn(self, bar, text, cmd, primary=False):
        b = tk.Button(bar, text=text, command=cmd, font=FONT_B if primary else FONT,
                      bg=C_BLUE if primary else "#FFFFFF",
                      fg="#FFFFFF" if primary else C_TEXT, relief="flat", bd=0,
                      padx=13, pady=6, cursor="hand2", highlightthickness=1,
                      highlightbackground=C_LINE)
        b.pack(side="left", padx=(0, 6))
        return b

    # =====================================================================
    # 데이터 · 그리기
    # =====================================================================
    def refresh(self):
        self.reload()

    def reload(self):
        self.nodes = self.db.flow_nodes()
        self.edges = self.db.flow_edges()
        self.npos = {}
        for i, n in enumerate(self.nodes):
            x = n["x"] if n["x"] is not None else 60 + (i % 3) * 300
            y = n["y"] if n["y"] is not None else 50 + (i // 3) * 160
            self.npos[n["id"]] = [x, y]
        if not self._fitted and self.nodes:
            self._fitted = True
            self.after(60, self.fit)          # 처음 열면 전체가 보이게
        self.draw()
        self.show_detail()

    # -- 확대 · 축소 ---------------------------------------------------------
    def content_size(self):
        """모델 좌표 기준 전체 크기."""
        if not self.npos:
            return 800, 600
        w = max(p[0] for p in self.npos.values()) + NODE_W
        h = max(p[1] for p in self.npos.values()) + NODE_H
        return w, h

    def set_scale(self, s):
        self.scale = max(0.30, min(2.0, s))
        self.v_zoom.set("%d%%" % round(self.scale * 100))
        self.draw()

    def zoom(self, factor):
        self.set_scale(self.scale * factor)

    def fit(self):
        """모든 항목이 화면에 들어오도록 배율을 맞춘다."""
        self.cv.update_idletasks()
        cw = max(self.cv.winfo_width(), 200)
        ch = max(self.cv.winfo_height(), 200)
        w, h = self.content_size()
        s = min((cw - 46) / float(w), (ch - 46) / float(h))
        self.set_scale(min(s, 1.3))
        self.cv.xview_moveto(0)
        self.cv.yview_moveto(0)
        self.app.set_status("화면에 맞춤 — %s (항목 %d개)"
                            % (self.v_zoom.get(), len(self.nodes)))

    def on_wheel(self, event):
        if event.state & 0x0004:                    # Ctrl + 휠 → 확대·축소
            self.zoom(1.1 if event.delta > 0 else 1 / 1.1)
        elif event.state & 0x0001:                  # Shift + 휠 → 좌우
            self.cv.xview_scroll(int(-event.delta / 120), "units")
        else:
            self.cv.yview_scroll(int(-event.delta / 120), "units")

    def on_canvas_resize(self, event=None):
        if self._fitted and self.nodes:
            self.after_idle(self.update_scroll)

    def toggle_detail(self):
        """오른쪽 상세를 접어 플로우차트를 넓게 본다."""
        if self.detail_on:
            self.pane.forget(self.right)
            self.b_detail.configure(text="상세 보이기", bg=C_GREEN, fg="#FFFFFF")
        else:
            self.pane.add(self.right, weight=4)
            self.b_detail.configure(text="상세 숨기기", bg="#FFFFFF", fg=C_TEXT)
        self.detail_on = not self.detail_on
        self.after(80, self.fit)

    # -- 그리기 --------------------------------------------------------------
    def draw(self):
        self.cv.delete("all")
        self.item_map = {}
        self.node_tag = {}
        if not self.nodes:
            self.cv.create_text(24, 24, anchor="nw", font=FONT, fill=C_MUTED,
                                text="플로우 항목이 없습니다. [＋ 항목 추가] 를 누르세요.")
            return
        self.draw_edges()
        counts = self.stage_counts()
        for n in self.nodes:
            self.draw_node(n, counts)
        self.update_scroll()

    def fnt(self, size, bold=False):
        n = max(6, int(round(size * self.scale)))
        return (F, n, "bold") if bold else (F, n)

    def update_scroll(self):
        w, h = self.content_size()
        s = self.scale
        cw = max(self.cv.winfo_width(), 100)
        ch = max(self.cv.winfo_height(), 100)
        self.cv.configure(scrollregion=(0, 0, max(w * s + 40, cw),
                                        max(h * s + 40, ch)))

    # -- 도형 위치 계산 ------------------------------------------------------
    def center(self, nid):
        x, y = self.npos[nid]
        return x + NODE_W / 2.0, y + NODE_H / 2.0

    def anchor(self, nid, other_id):
        """nid 도형의 테두리 위에서 other 쪽을 향하는 점."""
        cx, cy = self.center(nid)
        ox, oy = self.center(other_id)
        dx, dy = ox - cx, oy - cy
        if dx == 0 and dy == 0:
            return cx, cy
        w2, h2 = NODE_W / 2.0, NODE_H / 2.0
        shape = self.shape_of(nid)
        if shape == "diamond":
            k = 1.0 / (abs(dx) / w2 + abs(dy) / h2)
        elif shape == "oval":
            k = 1.0 / ((dx / w2) ** 2 + (dy / h2) ** 2) ** 0.5
        else:
            kx = w2 / abs(dx) if dx else float("inf")
            ky = h2 / abs(dy) if dy else float("inf")
            k = min(kx, ky)
        return cx + dx * k, cy + dy * k

    def shape_of(self, nid):
        for n in self.nodes:
            if n["id"] == nid:
                return (n["shape"] or "box")
        return "box"

    # -- 연결선 --------------------------------------------------------------
    def draw_edges(self):
        for it in self.cv.find_withtag("edge"):
            self.item_map.pop(it, None)
        self.cv.delete("edge")
        s = self.scale
        aw = max(6, int(12 * s)), max(7, int(15 * s)), max(3, int(5 * s))
        for e in self.edges:
            if e["src"] not in self.npos or e["dst"] not in self.npos:
                continue
            mx1, my1 = self.anchor(e["src"], e["dst"])
            mx2, my2 = self.anchor(e["dst"], e["src"])
            x1, y1, x2, y2 = mx1 * s, my1 * s, mx2 * s, my2 * s
            sel = (self.sel_kind == "edge" and self.sel_id == e["id"])
            it = self.cv.create_line(x1, y1, x2, y2, arrow="last",
                                     fill=C_BLUE if sel else "#93A2B4",
                                     width=3 if sel else 2,
                                     arrowshape=aw, tags=("edge",))
            self.item_map[it] = ("edge", e["id"])
            # 클릭 판정을 넉넉하게 (투명한 굵은 선)
            hit = self.cv.create_line(x1, y1, x2, y2, fill="", width=14,
                                      tags=("edge",))
            self.item_map[hit] = ("edge", e["id"])
            if e["label"]:
                mx, my = (x1 + x2) / 2.0, (y1 + y2) / 2.0
                t = self.cv.create_text(mx, my, text=e["label"], font=self.fnt(9),
                                        fill=C_TEXT, tags=("edge",))
                bb = self.cv.bbox(t)
                r = self.cv.create_rectangle(bb[0] - 4, bb[1] - 2, bb[2] + 4, bb[3] + 2,
                                             fill="#FAFBFD", outline="", tags=("edge",))
                self.cv.tag_raise(t, r)
                self.item_map[t] = ("edge", e["id"])
                self.item_map[r] = ("edge", e["id"])
        self.cv.tag_lower("edge")

    def draw_node(self, n, counts):
        nid = n["id"]
        s = self.scale
        x, y = self.npos[nid][0] * s, self.npos[nid][1] * s
        w, h = NODE_W * s, NODE_H * s
        tag = "nd%d" % nid
        self.node_tag[nid] = tag
        col = flow_color(n)
        sel = (self.sel_kind == "node" and self.sel_id == nid)
        conn = (self.connect_from == nid)
        outline = C_BLUE if sel else (C_GREEN if conn else C_LINE)
        width = 3 if (sel or conn) else 1
        bg = "#EAF1FD" if sel else ("#E6F6EC" if conn else "#FFFFFF")
        shape = n["shape"] or "box"
        detail = s >= 0.62          # 작게 줄이면 제목만 (한눈에 보기)

        if shape == "oval":
            body = self.cv.create_oval(x, y, x + w, y + h, fill=bg,
                                       outline=outline, width=width, tags=(tag,))
        elif shape == "diamond":
            body = self.cv.create_polygon(
                x + w / 2, y, x + w, y + h / 2, x + w / 2, y + h, x, y + h / 2,
                fill=bg, outline=outline, width=width, tags=(tag,))
        else:
            body = self.cv.create_rectangle(x, y, x + w, y + h, fill=bg,
                                            outline=outline, width=width, tags=(tag,))
        items = [body]
        if shape == "box":
            items.append(self.cv.create_rectangle(x, y, x + 6 * s, y + h, fill=col,
                                                  outline=col, tags=(tag,)))
            if detail:
                tx, ty, tw, anchor_ = x + 16 * s, y + 13 * s, w - 30 * s, "nw"
            else:
                tx, ty, tw, anchor_ = x + 14 * s, y + h / 2, w - 26 * s, "w"
        else:
            tx, tw, anchor_ = x + w / 2, w - 46 * s, "center"
            ty = y + h / 2 if not detail else y + 18 * s
            if detail:
                anchor_ = "n"

        items.append(self.cv.create_text(
            tx, ty, anchor=anchor_, width=tw, text=n["title"] or "(이름 없음)",
            font=self.fnt(10, True), fill=col if shape != "box" else C_TEXT,
            justify="center" if shape != "box" else "left", tags=(tag,)))

        if shape == "box" and detail:
            sub = (n["descr"] or "").replace("\n", " ")
            if len(sub) > 58:
                sub = sub[:58] + "…"
            items.append(self.cv.create_text(x + 16 * s, y + 38 * s, anchor="nw",
                                             width=w - 30 * s, text=sub,
                                             font=self.fnt(9), fill=C_MUTED, tags=(tag,)))
            if n["store"]:
                items.append(self.cv.create_text(
                    x + 16 * s, y + h - 17 * s, anchor="nw", width=w - 30 * s,
                    text="📁 " + n["store"], font=self.fnt(8), fill="#8A96A4",
                    tags=(tag,)))

        cnt = counts.get(n["stage_no"], 0) if n["stage_no"] is not None else 0
        if cnt:
            r = 11 * s
            cx, cy = x + w - 21 * s, y + 19 * s
            items.append(self.cv.create_oval(cx - r, cy - r, cx + r, cy + r,
                                             fill=col, outline=col, tags=(tag,)))
            items.append(self.cv.create_text(cx, cy, text=str(cnt),
                                             font=self.fnt(9, True), fill="#FFFFFF",
                                             tags=(tag,)))
        for it in items:
            self.item_map[it] = ("node", nid)

    def stage_counts(self):
        out = {}
        for r in self.db.all_deals():
            if (r["status"] or "진행중") != "진행중":
                continue
            _n, nxt = self.db.progress(r["id"])
            if nxt is not None:
                out[nxt] = out.get(nxt, 0) + 1
        return out

    # =====================================================================
    # 마우스
    # =====================================================================
    def hit(self, event):
        x, y = self.cv.canvasx(event.x), self.cv.canvasy(event.y)
        for it in reversed(self.cv.find_overlapping(x - 2, y - 2, x + 2, y + 2)):
            if it in self.item_map:
                return self.item_map[it], (x, y)
        return (None, None), (x, y)

    def on_press(self, event):
        self.cv.focus_set()
        (kind, oid), (x, y) = self.hit(event)
        if self.connect_mode:
            if kind == "node":
                if self.connect_from is None:
                    self.connect_from = oid
                    self.v_hint.set("연결 그리기 : 도착할 항목을 클릭하세요 (Esc 취소)")
                    self.draw()
                elif oid != self.connect_from:
                    self.db.add_edge(self.connect_from, oid)
                    self.connect_from = None
                    self.v_hint.set("연결 그리기 : 시작할 항목을 클릭하세요 (Esc 취소)")
                    self.edges = self.db.flow_edges()
                    self.draw()
                    self.app.set_status("연결선을 추가했습니다.")
            else:
                self.cancel_connect()
            return

        if kind == "node":
            self.select("node", oid, redraw=False)
            self._drag = {"id": oid, "x": x, "y": y, "moved": False}
            self.draw()
        elif kind == "edge":
            self.select("edge", oid)
        else:
            self.select(None, None)

    def on_drag(self, event):
        if not self._drag:
            return
        x, y = self.cv.canvasx(event.x), self.cv.canvasy(event.y)
        dx, dy = x - self._drag["x"], y - self._drag["y"]
        if not dx and not dy:
            return
        s = self.scale
        nid = self._drag["id"]
        pos = self.npos[nid]
        nx, ny = max(0, pos[0] + dx / s), max(0, pos[1] + dy / s)
        self.cv.move(self.node_tag[nid], (nx - pos[0]) * s, (ny - pos[1]) * s)
        self.npos[nid] = [nx, ny]
        self._drag["x"], self._drag["y"] = x, y
        self._drag["moved"] = True
        self.draw_edges()

    def on_release(self, event):
        if not self._drag:
            return
        nid = self._drag["id"]
        if self._drag["moved"]:
            x, y = self.npos[nid]
            x = int(round(x / SNAP) * SNAP)
            y = int(round(y / SNAP) * SNAP)
            self.npos[nid] = [x, y]
            self.db.set_flow_pos(nid, x, y)
            self.nodes = self.db.flow_nodes()
            self.draw()
            self.update_scroll()
            self.app.set_status("위치 저장 (%d, %d)" % (x, y))
        self._drag = None

    def on_double(self, event):
        (kind, oid), _p = self.hit(event)
        if kind == "node":
            self.select("node", oid)
            self.edit_node()
        elif kind == "edge":
            self.select("edge", oid)
            self.edit_edge_label(oid)

    # =====================================================================
    # 선택 · 편집
    # =====================================================================
    def select(self, kind, oid, redraw=True):
        self.sel_kind, self.sel_id = kind, oid
        if redraw:
            self.draw()
        self.show_detail()

    def current_node(self):
        if self.sel_kind == "node" and self.sel_id:
            return self.db.flow_node(self.sel_id)
        return None

    def free_spot(self):
        """겹치지 않는 빈 자리를 찾아준다."""
        used = [(p[0], p[1]) for p in self.npos.values()]
        for r in range(0, 40):
            for c in range(0, 4):
                x, y = 60 + c * 300, 50 + r * 160
                if all(abs(x - ux) > 40 or abs(y - uy) > 40 for ux, uy in used):
                    return x, y
        return 60, 50

    def toggle_connect(self):
        self.connect_mode = not self.connect_mode
        self.connect_from = None
        if self.connect_mode:
            self.b_conn.configure(bg=C_GREEN, fg="#FFFFFF", font=FONT_B)
            self.v_hint.set("연결 그리기 : 시작할 항목을 클릭하세요 (Esc 취소)")
        else:
            self.b_conn.configure(bg="#FFFFFF", fg=C_TEXT, font=FONT)
            self.v_hint.set("")
        self.draw()

    def cancel_connect(self):
        if self.connect_mode:
            self.connect_mode = False
            self.connect_from = None
            self.b_conn.configure(bg="#FFFFFF", fg=C_TEXT, font=FONT)
            self.v_hint.set("")
            self.draw()

    def add_node(self):
        FlowEditor(self)

    def edit_node(self):
        n = self.current_node()
        if not n:
            messagebox.showinfo(APP_TITLE, "수정할 항목을 클릭해서 고르세요.", parent=self)
            return
        FlowEditor(self, n)

    def delete_sel(self):
        if self.sel_kind == "edge" and self.sel_id:
            if messagebox.askyesno(APP_TITLE, "선택한 연결선을 삭제할까요?", parent=self):
                self.db.delete_edge(self.sel_id)
                self.select(None, None)
                self.reload()
                self.app.set_status("연결선 삭제")
            return
        n = self.current_node()
        if not n:
            messagebox.showinfo(APP_TITLE,
                                "삭제할 항목이나 연결선을 클릭해서 고르세요.", parent=self)
            return
        if messagebox.askyesno(APP_TITLE,
                               "'%s' 항목과 연결선을 삭제할까요?\n"
                               "(안건 데이터와 자료방 파일은 그대로 남습니다)" % n["title"],
                               parent=self):
            self.db.delete_flow(n["id"])
            self.select(None, None)
            self.reload()
            self.app.set_status("플로우 항목 삭제")

    def edit_edge_label(self, eid):
        e = [x for x in self.edges if x["id"] == eid]
        if not e:
            return
        win = tk.Toplevel(self)
        win.title("연결선 설명")
        win.configure(bg=C_BG)
        win.transient(self.winfo_toplevel())
        win.resizable(False, False)
        win.grab_set()
        tk.Label(win, text="연결선에 표시할 글자 (예: 승인, 반려, 재견적)",
                 font=FONT, bg=C_BG, fg=C_TEXT).pack(anchor="w", padx=20, pady=(18, 6))
        v = tk.StringVar(value=e[0]["label"] or "")
        ent = ttk.Entry(win, textvariable=v, width=40, font=FONT)
        ent.pack(padx=20)
        ent.focus_set()

        def ok():
            self.db.set_edge_label(eid, v.get().strip())
            win.destroy()
            self.reload()
            self.select("edge", eid)
        ent.bind("<Return>", lambda ev: ok())
        row = tk.Frame(win, bg=C_BG)
        row.pack(fill="x", padx=20, pady=(12, 18))
        tk.Button(row, text="저장", command=ok, font=FONT_B, bg=C_BLUE, fg="#FFFFFF",
                  relief="flat", bd=0, padx=20, pady=6, cursor="hand2").pack(side="right")
        tk.Button(row, text="취소", command=win.destroy, font=FONT, bg="#FFFFFF",
                  fg=C_TEXT, relief="flat", bd=0, padx=18, pady=6, cursor="hand2",
                  highlightthickness=1, highlightbackground=C_LINE
                  ).pack(side="right", padx=(0, 8))
        win.update_idletasks()
        win.geometry("+%d+%d" % (self.winfo_rootx() + 240, self.winfo_rooty() + 200))

    def auto_layout(self):
        if messagebox.askyesno(APP_TITLE,
                               "모든 항목을 순서대로 격자 배치합니다.\n"
                               "직접 옮긴 위치는 사라집니다. 진행할까요?", parent=self):
            self.db.auto_layout()
            self.reload()
            self.app.set_status("자동 정렬 완료")

    def reset(self):
        if messagebox.askyesno(APP_TITLE,
                               "플로우를 엑셀 기준 0~8단계로 되돌립니다.\n"
                               "직접 추가·수정한 항목과 연결선은 사라집니다. 진행할까요?",
                               parent=self):
            self.db.reset_flow()
            self.select(None, None)
            self.reload()
            self.app.set_status("플로우를 기본값으로 되돌렸습니다.")

    # =====================================================================
    # 오른쪽 상세
    # =====================================================================
    def archive_root(self):
        return self.db.get_setting("archive_path", DEFAULT_SETTINGS["archive_path"])

    def node_folder(self, n):
        root = self.archive_root()
        f = (n["folder"] or "").strip()
        if f:
            p = f if os.path.isabs(f) else os.path.join(root, f)
            if os.path.isdir(p):
                return p
        sn = n["stage_no"]
        if sn is not None and 0 <= sn < STAGE_COUNT and os.path.isdir(root):
            _sub, kw = STAGE_FOLDER.get(sn, (None, None))
            if kw:
                for name in sorted(os.listdir(root)):
                    if kw in name and os.path.isdir(os.path.join(root, name)):
                        return os.path.join(root, name)
        return root if os.path.isdir(root) else None

    def show_detail(self):
        n = self.current_node()
        self.txt.configure(state="normal")
        self.txt.delete("1.0", "end")
        self.file_rows = {}
        if not n:
            if self.sel_kind == "edge":
                e = [x for x in self.edges if x["id"] == self.sel_id]
                if e:
                    src = self.db.flow_node(e[0]["src"])
                    dst = self.db.flow_node(e[0]["dst"])
                    self.v_ntitle.set("연결선")
                    self.txt.insert("1.0", "%s\n    ↓  %s\n%s\n\n"
                                           "· 더블클릭하면 연결선에 글자를 넣을 수 있습니다\n"
                                           "· [삭제] 를 누르면 이 연결선이 지워집니다"
                                    % (src["title"] if src else "?",
                                       e[0]["label"] or "",
                                       dst["title"] if dst else "?"))
            else:
                self.v_ntitle.set("항목을 클릭하세요")
                self.txt.insert("1.0",
                                "· 박스를 끌어서 원하는 위치에 놓으세요\n"
                                "· [🔗 연결 그리기] 로 항목끼리 화살표를 잇습니다\n"
                                "· 박스를 더블클릭하면 내용을 고칠 수 있습니다\n"
                                "· 항목을 클릭하면 관련 자료와 안건이 여기에 나옵니다")
            self.v_folder.set("")
            self.txt.configure(state="disabled")
            self.g_files.set_rows([])
            self.g_deals.set_rows([])
            return

        self.v_ntitle.set(n["title"] or "(이름 없음)")
        if n["descr"]:
            self.txt.insert("end", n["descr"] + "\n")
        if n["tip"]:
            self.txt.insert("end", "\n✔  " + n["tip"] + "\n", "tip")
        if n["store"]:
            self.txt.insert("end", "\n저장소 : %s\n" % n["store"], "store")
        self.txt.configure(state="disabled")

        folder = self.node_folder(n)
        root = self.archive_root()
        if folder:
            try:
                rel = os.path.relpath(folder, os.path.dirname(root))
            except ValueError:
                rel = folder
            self.v_folder.set("📁  " + rel.replace(os.sep, " › "))
        else:
            self.v_folder.set("자료방 폴더를 찾을 수 없습니다")

        rows = []
        if folder:
            try:
                names = sorted(os.listdir(folder))
            except OSError:
                names = []
            dirs = [x for x in names if os.path.isdir(os.path.join(folder, x))]
            for i, name in enumerate(dirs + [x for x in names if x not in dirs]):
                full = os.path.join(folder, name)
                iid = "f%d" % i
                self.file_rows[iid] = full
                if name in dirs:
                    try:
                        cnt = len(os.listdir(full))
                    except OSError:
                        cnt = 0
                    rows.append((iid, ["📁  " + name, "폴더", "%d개" % cnt,
                                       mtime(full)], "ok"))
                else:
                    try:
                        sz = os.path.getsize(full)
                    except OSError:
                        sz = 0
                    rows.append((iid, ["      " + name, file_kind(name), fsize(sz),
                                       mtime(full)], ""))
        self.g_files.set_rows(rows)

        drows = []
        sn = n["stage_no"]
        if sn is not None and sn >= 0:
            for r in self.db.all_deals():
                if (r["status"] or "진행중") != "진행중":
                    continue
                _c, nxt = self.db.progress(r["id"])
                if nxt != sn:
                    continue
                label, _d, days = self.db.next_deadline(r)
                drows.append((str(r["id"]), [
                    r["code"] or "", r["customer"] or "", r["model"] or "",
                    "%s %s" % (label, dday_text(days)) if label else ""],
                    _tag_by_days(days)))
        self.g_deals.set_rows(drows)

    def open_file(self, iid):
        path = self.file_rows.get(iid)
        if not path:
            return
        if os.path.isdir(path):
            self.app.show("archive")
            self.app.pages["archive"].goto(path)
        else:
            open_path(path, self)

    def open_in_archive(self):
        n = self.current_node()
        folder = self.node_folder(n) if n else self.archive_root()
        if not folder or not os.path.isdir(folder):
            messagebox.showinfo(APP_TITLE, "자료방 폴더를 찾을 수 없습니다.\n"
                                           "[도구 → 설정] 에서 경로를 확인하세요.",
                                parent=self)
            return
        self.app.show("archive")
        self.app.pages["archive"].goto(folder)


# ===========================================================================
# 메인 윈도우
# ===========================================================================

# 좌측 메뉴
MENU = [
    ("dash",    "대시보드",    DashboardPage),
    ("deals",   "안건 관리",   DealsPage),
    (None,      None,         None),          # 구분선
    ("archive", "자료방",      ArchivePage),
    ("flow",    "업무 플로우", FlowPage),
]


class App(tk.Tk):
    def __init__(self):
        tk.Tk.__init__(self)
        self.title(APP_TITLE)
        self.geometry("1500x900")
        self.minsize(1240, 760)
        self.configure(bg=C_BG)

        self.db = Database(DB_PATH)
        self.pages = {}
        self.menu_items = {}
        self.current = None

        self._style()
        self._menubar()
        self._header()
        self._layout()
        self.show("dash")
        self.protocol("WM_DELETE_WINDOW", self.on_close)
        self.bind("<Control-n>", lambda e: self._deals_action("new_deal"))
        self.bind("<Control-s>", lambda e: self._deals_action("save_deal"))
        self.bind("<F5>", lambda e: self.refresh_all())

    # -- 뼈대 ---------------------------------------------------------------
    def _style(self):
        st = ttk.Style(self)
        try:
            st.theme_use("clam")
        except tk.TclError:
            pass
        st.configure(".", font=FONT, background=C_BG, foreground=C_TEXT)
        st.configure("Treeview", font=FONT, rowheight=27, fieldbackground=C_CARD,
                     background=C_CARD, borderwidth=0)
        st.configure("Treeview.Heading", font=FONT_SB, background="#E3E9F0",
                     foreground=C_TEXT, relief="flat", padding=(6, 6))
        st.map("Treeview.Heading", background=[("active", "#D5DEE9")])
        st.map("Treeview", background=[("selected", "#CBDDF7")],
               foreground=[("selected", C_TEXT)])
        st.configure("TNotebook", background=C_BG, borderwidth=0)
        st.configure("TNotebook.Tab", font=FONT, padding=(16, 7),
                     background="#DCE3EC", foreground=C_MUTED)
        st.map("TNotebook.Tab", background=[("selected", C_CARD)],
               foreground=[("selected", C_BLUE)])
        st.configure("TCombobox", fieldbackground=C_CARD)
        st.configure("TEntry", fieldbackground=C_CARD)
        st.configure("TPanedwindow", background=C_BG)

    def _menubar(self):
        m = tk.Menu(self)
        f = tk.Menu(m, tearoff=0, font=FONT)
        f.add_command(label="새 안건", accelerator="Ctrl+N",
                      command=lambda: self._deals_action("new_deal"))
        f.add_command(label="저장", accelerator="Ctrl+S",
                      command=lambda: self._deals_action("save_deal"))
        f.add_separator()
        f.add_command(label="엑셀로 내보내기...", command=self.export_excel)
        f.add_command(label="데이터 백업...", command=self.backup_db)
        f.add_separator()
        f.add_command(label="종료", command=self.on_close)
        m.add_cascade(label="파일", menu=f)

        v = tk.Menu(m, tearoff=0, font=FONT)
        for key, label, _cls in MENU:
            if key:
                v.add_command(label=label, command=lambda k=key: self.show(k))
        v.add_separator()
        v.add_command(label="새로고침", accelerator="F5", command=self.refresh_all)
        m.add_cascade(label="화면", menu=v)

        t = tk.Menu(m, tearoff=0, font=FONT)
        t.add_command(label="설정 · 기준정보...", command=self.open_settings)
        m.add_cascade(label="도구", menu=t)

        h = tk.Menu(m, tearoff=0, font=FONT)
        h.add_command(label="프로그램 정보", command=self.about)
        m.add_cascade(label="도움말", menu=h)
        self.config(menu=m)

    def _header(self):
        h = tk.Frame(self, bg=C_HEADER, height=58)
        h.pack(fill="x")
        h.pack_propagate(False)
        tk.Label(h, text="DSS 수출입 관리", font=(F, 15, "bold"), bg=C_HEADER,
                 fg="#FFFFFF").pack(side="left", padx=(22, 10))
        tk.Label(h, text="Import / Export Management", font=FONT_S, bg=C_HEADER,
                 fg=C_HEADER_SUB).pack(side="left", pady=(6, 0))
        self.v_head = tk.StringVar()
        tk.Label(h, textvariable=self.v_head, font=FONT, bg=C_HEADER,
                 fg="#DCE6F4").pack(side="right", padx=22)
        self.refresh_header()

    def refresh_header(self):
        # 담당자명은 표시하지 않는다 (Offer Sheet 메일 발신자에만 사용)
        d = datetime.date.today()
        wd = "월화수목금토일"[d.weekday()]
        self.v_head.set("%s   ·   %04d-%02d-%02d (%s)" % (
            self.db.get_setting("company", "㈜디에스에스"), d.year, d.month, d.day, wd))

    def _layout(self):
        body = tk.Frame(self, bg=C_BG)
        body.pack(fill="both", expand=True)

        side = tk.Frame(body, bg=C_SIDE, width=190)
        side.pack(side="left", fill="y")
        side.pack_propagate(False)
        tk.Frame(side, bg=C_SIDE, height=10).pack(fill="x")

        for key, label, _cls in MENU:
            if key is None:
                tk.Frame(side, bg="#33465E", height=1).pack(fill="x", padx=16, pady=10)
                continue
            item = tk.Label(side, text="   " + label, font=FONT, bg=C_SIDE,
                            fg=C_SIDE_TEXT, anchor="w", padx=14, pady=11,
                            cursor="hand2")
            item.pack(fill="x")
            item.bind("<Button-1>", lambda e, k=key: self.show(k))
            item.bind("<Enter>", lambda e, k=key: self._hover(k, True))
            item.bind("<Leave>", lambda e, k=key: self._hover(k, False))
            self.menu_items[key] = item

        self.body = tk.Frame(body, bg=C_BG)
        self.body.pack(side="left", fill="both", expand=True)

        sb = tk.Frame(self, bg="#DDE3EA", height=26)
        sb.pack(fill="x")
        sb.pack_propagate(False)
        self.v_status = tk.StringVar(value="준비됨")
        tk.Label(sb, textvariable=self.v_status, bg="#DDE3EA", fg=C_MUTED,
                 font=FONT_S, anchor="w").pack(side="left", padx=14)
        tk.Label(sb, text="데이터: %s" % DB_PATH, bg="#DDE3EA", fg="#94A0AE",
                 font=FONT_S).pack(side="right", padx=14)

    def _hover(self, key, on):
        if key == self.current:
            return
        self.menu_items[key].configure(bg=C_SIDE_HOVER if on else C_SIDE)

    # -- 페이지 -------------------------------------------------------------
    def show(self, key):
        if key == self.current:
            self.pages[key].refresh()
            return
        if self.current == "deals" and "deals" in self.pages:
            if not self.pages["deals"].confirm_discard():
                return
        if key not in self.pages:
            cls = dict((k, c) for k, _l, c in MENU if k)[key]
            self.pages[key] = cls(self)
        for k, p in self.pages.items():
            p.pack_forget()
        for k, item in self.menu_items.items():
            on = (k == key)
            item.configure(bg=C_SIDE_ACTIVE if on else C_SIDE,
                           fg=C_SIDE_TEXT_ON if on else C_SIDE_TEXT,
                           font=FONT_B if on else FONT)
        self.current = key
        self.pages[key].pack(fill="both", expand=True)
        self.pages[key].refresh()

    def refresh_all(self):
        for p in self.pages.values():
            p.refresh()
        self.refresh_header()
        self.set_status("새로고침 완료")

    def refresh_others(self, except_key):
        for k, p in self.pages.items():
            if k != except_key:
                try:
                    p.refresh()
                except Exception:
                    pass

    def open_deal(self, deal_id):
        self.show("deals")
        self.pages["deals"].open_deal(deal_id)

    def open_deals_view(self, view):
        self.show("deals")
        self.pages["deals"].show_view(view)

    def _deals_action(self, name):
        self.show("deals")
        getattr(self.pages["deals"], name)()

    def set_status(self, msg):
        self.v_status.set(msg)

    # -- 공용 창 ------------------------------------------------------------
    def open_settings(self):
        """기준정보(선택목록 · 신용장 규칙 · 기본값)를 별도 창으로 연다."""
        win = tk.Toplevel(self)
        win.title("설정 · 기준정보")
        win.geometry("1020x740")
        win.configure(bg=C_BG)
        win.transient(self)
        page = MasterPage(self, parent=win)
        page.pack(fill="both", expand=True)
        page.refresh()
        bar = tk.Frame(win, bg=C_BG)
        bar.pack(fill="x", padx=22, pady=(0, 14))
        tk.Button(bar, text="닫기", command=win.destroy, font=FONT, relief="flat",
                  bd=0, bg="#FFFFFF", fg=C_TEXT, padx=20, pady=7, cursor="hand2",
                  highlightthickness=1, highlightbackground=C_LINE).pack(side="right")

    def show_text_window(self, title, text):
        win = tk.Toplevel(self)
        win.title(title)
        win.geometry("900x620")
        win.configure(bg=C_BG)
        win.transient(self)
        tk.Label(win, text=title, font=FONT_H, bg=C_BG, fg=C_TEXT
                 ).pack(anchor="w", padx=18, pady=(16, 2))
        tk.Label(win, text="내용을 수정한 뒤 복사할 수 있습니다.", font=FONT_S,
                 bg=C_BG, fg=C_MUTED).pack(anchor="w", padx=18)
        t = tk.Text(win, font=FONT, wrap="none", padx=14, pady=12, relief="flat",
                    bg=C_CARD, highlightthickness=1, highlightbackground=C_LINE)
        t.pack(fill="both", expand=True, padx=18, pady=(10, 4))
        t.insert("1.0", text)
        hb = ttk.Scrollbar(win, orient="horizontal", command=t.xview)
        t.configure(xscrollcommand=hb.set)
        hb.pack(fill="x", padx=18)

        bar = tk.Frame(win, bg=C_BG)
        bar.pack(fill="x", padx=18, pady=12)

        def copy():
            self.clipboard_clear()
            self.clipboard_append(t.get("1.0", "end").rstrip())
            self.set_status("클립보드에 복사했습니다.")
            win.destroy()

        tk.Button(bar, text="클립보드로 복사", command=copy, font=FONT_B, relief="flat",
                  bd=0, bg=C_BLUE, fg="#FFFFFF", padx=20, pady=7, cursor="hand2"
                  ).pack(side="right")
        tk.Button(bar, text="닫기", command=win.destroy, font=FONT, relief="flat",
                  bd=0, bg="#FFFFFF", fg=C_TEXT, padx=18, pady=7, cursor="hand2",
                  highlightthickness=1, highlightbackground=C_LINE
                  ).pack(side="right", padx=(0, 8))

    # -- 내보내기 / 백업 -----------------------------------------------------
    def export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror(APP_TITLE,
                                 "엑셀 내보내기에는 openpyxl 이 필요합니다.\n\n"
                                 "명령 프롬프트에서 아래를 실행하세요:\n    pip install openpyxl",
                                 parent=self)
            return
        path = filedialog.asksaveasfilename(
            parent=self, title="엑셀로 내보내기", defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            initialfile="수출입안건_%s.xlsx" % datetime.date.today().strftime("%Y%m%d"))
        if not path:
            return

        wb = openpyxl.Workbook()
        thin = Border(*[Side(style="thin", color="D0D0D0")] * 4)
        hf = PatternFill("solid", fgColor="1B2F52")
        hfont = Font(name=F, size=10, bold=True, color="FFFFFF")
        bfont = Font(name=F, size=10)

        def style_head(ws):
            for c in ws[1]:
                c.fill, c.font, c.border = hf, hfont, thin
                c.alignment = Alignment(horizontal="center", vertical="center",
                                        wrap_text=True)

        # 1) 안건목록
        ws = wb.active
        ws.title = "안건목록"
        heads = ["안건번호", "상태", "고객사", "담당자", "최종고객사", "사이트",
                 "모델/품명", "수량", "구분", "P.O 번호", "P.O 접수일", "Debit Note",
                 "Offer Sheet", "FCA", "Latest shipment", "Expiry date", "L/C 번호",
                 "개설은행", "금액", "통화", "커미션 요율", "커미션 금액",
                 "청구일", "수령일", "다음 기한", "진행단계", "진척", "비고"]
        keys = ["code", "status", "customer", "customer_pic", "end_user", "site",
                "model", "qty", "kind", "po_no", "po_date", "debit_no", "offer_no",
                "fca_date", "latest_shipment", "expiry_date", "lc_no", "lc_bank",
                "amount", "currency", "comm_rate", "comm_amount",
                "comm_billed", "comm_received"]
        ws.append(heads)
        style_head(ws)
        for r in self.db.all_deals():
            n_done, nxt = self.db.progress(r["id"])
            label, date_s, days = self.db.next_deadline(r)
            ws.append([r[k] or "" for k in keys] + [
                "%s %s (%s)" % (label, dday_text(days), date_s) if label else "",
                "완료" if nxt is None else "%d. %s" % (nxt, STAGE_NAME[nxt]),
                "%d/%d" % (n_done, STAGE_COUNT), r["note"] or ""])
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.font, c.border = bfont, thin
        for i, w in enumerate([11, 8, 15, 10, 12, 14, 22, 10, 8, 15, 12, 14, 14,
                               12, 14, 14, 18, 12, 14, 7, 10, 14, 12, 12, 26, 22,
                               8, 30], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # 2) 진행현황
        ws2 = wb.create_sheet("진행현황")
        ws2.append(["안건번호", "고객사", "P.O 번호"] +
                   ["%d. %s" % (n, STAGE_NAME[n]) for n in range(STAGE_COUNT)] + ["진척"])
        style_head(ws2)
        okf = PatternFill("solid", fgColor="DDF3DD")
        for r in self.db.all_deals():
            st = self.db.stages(r["id"])
            n_done, _n = self.db.progress(r["id"])
            ws2.append([r["code"] or "", r["customer"] or "", r["po_no"] or ""] +
                       [st.get(n, ("", ""))[0] or "" for n in range(STAGE_COUNT)] +
                       ["%d/%d" % (n_done, STAGE_COUNT)])
            for i, c in enumerate(ws2[ws2.max_row]):
                c.font, c.border = bfont, thin
                c.alignment = Alignment(horizontal="center", vertical="center")
                if 3 <= i < 3 + STAGE_COUNT and c.value:
                    c.fill = okf
        for i, w in enumerate([11, 15, 15] + [15] * STAGE_COUNT + [8], 1):
            ws2.column_dimensions[get_column_letter(i)].width = w
        ws2.freeze_panes = "D2"

        # 3) 신용장현황
        ws3 = wb.create_sheet("신용장현황")
        ws3.append(["안건번호", "고객사", "P.O 번호", "FCA", "개설요청 기한",
                    "개설은행", "L/C 번호", "개설일", "Latest shipment",
                    "Expiry date", "적용규칙", "Amend"])
        style_head(ws3)
        lead = self.db.get_int("lc_lead", 7)
        for r in self.db.all_deals():
            ship, exp = self.db.lc_rule(r["customer"])
            ws3.append([r["code"] or "", r["customer"] or "", r["po_no"] or "",
                        norm_date(r["fca_date"]), add_days(r["fca_date"], -lead),
                        r["lc_bank"] or "", r["lc_no"] or "",
                        norm_date(r["lc_open_date"]), norm_date(r["latest_shipment"]),
                        norm_date(r["expiry_date"]), "+%d일 / +%d일" % (ship, exp),
                        r["lc_amend"] or ""])
        for row in ws3.iter_rows(min_row=2):
            for c in row:
                c.font, c.border = bfont, thin
        for i, w in enumerate([11, 15, 15, 12, 13, 12, 20, 12, 15, 13, 14, 20], 1):
            ws3.column_dimensions[get_column_letter(i)].width = w
        ws3.freeze_panes = "A2"

        # 4) 커미션현황
        ws4 = wb.create_sheet("커미션현황")
        ws4.append(["안건번호", "고객사", "P.O 번호", "납품금액", "통화", "요율(%)",
                    "커미션 금액", "청구서 번호", "청구일", "수령일", "상태"])
        style_head(ws4)
        due = self.db.get_int("comm_due", 30)
        for r in self.db.all_deals():
            b, v = parse_date(r["comm_billed"]), parse_date(r["comm_received"])
            if v:
                state = "수령 완료"
            elif b:
                n = (datetime.date.today() - b).days
                state = "입금 지연 (%d일)" % n if n >= due else "입금 대기 (%d일)" % n
            elif 7 in self.db.done_set(r["id"]):
                state = "청구 필요"
            else:
                state = ""
            ws4.append([r["code"] or "", r["customer"] or "", r["po_no"] or "",
                        money(r["amount"]), r["currency"] or "", r["comm_rate"] or "",
                        money(r["comm_amount"]), r["comm_invoice_no"] or "",
                        norm_date(r["comm_billed"]), norm_date(r["comm_received"]), state])
        for row in ws4.iter_rows(min_row=2):
            for c in row:
                c.font, c.border = bfont, thin
        for i, w in enumerate([11, 15, 15, 14, 7, 8, 14, 30, 12, 12, 16], 1):
            ws4.column_dimensions[get_column_letter(i)].width = w
        ws4.freeze_panes = "A2"

        # 5) 기한알림
        ws5 = wb.create_sheet("기한알림")
        ws5.append(["구분", "D-Day", "기준일", "항목", "안건번호", "고객사", "P.O", "내용"])
        style_head(ws5)
        for a in compute_alerts(self.db):
            ws5.append([a["level"], dday_text(a["dday"]), a["date"], a["item"],
                        a["code"], a["customer"], a["po"], a["detail"]])
        for row in ws5.iter_rows(min_row=2):
            for c in row:
                c.font, c.border = bfont, thin
            if row[0].value == "지연":
                for c in row:
                    c.fill = PatternFill("solid", fgColor="FFD9D9")
            elif row[0].value == "임박":
                for c in row:
                    c.fill = PatternFill("solid", fgColor="FFF2CC")
        for i, w in enumerate([8, 9, 12, 20, 11, 15, 16, 60], 1):
            ws5.column_dimensions[get_column_letter(i)].width = w
        ws5.freeze_panes = "A2"

        try:
            wb.save(path)
        except IOError:
            messagebox.showerror(APP_TITLE, "파일을 저장할 수 없습니다.\n"
                                            "같은 이름의 엑셀이 열려 있는지 확인하세요.",
                                 parent=self)
            return
        self.set_status("엑셀 내보내기 완료: %s" % path)
        if messagebox.askyesno(APP_TITLE, "내보내기가 완료되었습니다.\n지금 열어볼까요?",
                               parent=self):
            try:
                os.startfile(path)
            except Exception:
                pass

    def backup_db(self):
        path = filedialog.asksaveasfilename(
            parent=self, title="데이터 백업", defaultextension=".db",
            filetypes=[("데이터베이스", "*.db")],
            initialfile="dss_trade_backup_%s.db" % datetime.date.today().strftime("%Y%m%d"))
        if not path:
            return
        self.db.con.commit()
        shutil.copyfile(DB_PATH, path)
        self.set_status("백업 완료: %s" % path)

    def about(self):
        messagebox.showinfo(APP_TITLE,
                            "%s\n\n"
                            "「1. 수출입.영업.관리 업무 정리」 엑셀의 외자(수출입) 업무를\n"
                            "안건 단위로 관리하는 프로그램입니다.\n\n"
                            "· 영업_외자 업무 흐름 0~8단계 진행 추적\n"
                            "· FCA 기준 신용장 기한 자동 산정 (고객사별 규칙)\n"
                            "· 발주·납기 / 신용장 / 선적·서류 / 커미션 현황\n"
                            "· 기한 임박·경과 자동 알림\n"
                            "· Offer Sheet 송부 메일 문안 생성\n"
                            "· 엑셀 내보내기 (5개 시트)\n\n"
                            "데이터 저장 위치\n%s" % (APP_TITLE, DB_PATH), parent=self)

    def on_close(self):
        if "deals" in self.pages and not self.pages["deals"].confirm_discard():
            return
        try:
            self.db.con.close()
        except Exception:
            pass
        self.destroy()


def main():
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass
    App().mainloop()


if __name__ == "__main__":
    main()